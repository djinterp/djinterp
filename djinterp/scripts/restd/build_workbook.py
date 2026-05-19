"""
restd coverage workbook builder.

Reads one JSON file per std header from <root>/docs/restd/data/<header>.json
and writes the workbook to <root>/docs/restd/restd_coverage.xlsx.

Project layout (paths relative to this script's location, scripts/restd/):
    ../../docs/restd/data/        -- per-header JSON files (source of truth)
    ../../docs/restd/data/_roadmap.json
    ../../docs/restd/restd_coverage.xlsx   (output)

Run:  python3 scripts/restd/build_workbook.py
  (or from anywhere — paths are resolved relative to this file via __file__)

The JSON layer is the single source of truth. The constants in this
file (Legend column-meanings, the aux-sheet blurbs) change rarely
and stay hardcoded — they describe the workbook itself, not its
data. To add a new header: drop a new file into docs/restd/data/
and rerun. To change the column structure: edit COLUMNS below.
"""

import glob
import json
import os
from collections import namedtuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Project paths. Resolved relative to this script's location so the
# builder runs identically from any cwd.
#   scripts/restd/build_workbook.py     -> __file__ dir is scripts/restd/
#   ../../docs/restd/data/              -> JSON source of truth
#   ../../docs/restd/restd_coverage.xlsx -> workbook output
# ---------------------------------------------------------------------------

_HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.normpath(os.path.join(_HERE, "..", ".."))
DATA_DIR = os.path.join(PROJECT_ROOT, "docs", "restd", "data")
OUTPUT_PATH = os.path.join(PROJECT_ROOT, "docs", "restd", "restd_coverage.xlsx")

# ---------------------------------------------------------------------------
# Symbol record — same shape as the old coverage_data.Symbol, so the
# existing sheet-building logic doesn't need to change.
# ---------------------------------------------------------------------------

Symbol = namedtuple("Symbol", [
    "name", "group",
    "std_in", "restd_min",
    "constexpr_in_std", "constexpr_in_restd",
    "intrinsic_required", "intrinsic_names", "detection_macro",
    "t_alias_in", "v_var_in", "deprecated_in",
    "notes", "depends_on", "failure_reason",
])


def symbol_from_dict(d):
    """JSON dict -> Symbol namedtuple. Tolerates missing optional fields."""
    return Symbol(
        name=d["name"],
        group=d["group"],
        std_in=d.get("std_in"),
        restd_min=d.get("restd_min"),
        constexpr_in_std=d.get("constexpr_in_std"),
        constexpr_in_restd=d.get("constexpr_in_restd"),
        intrinsic_required=bool(d.get("intrinsic_required", False)),
        intrinsic_names=d.get("intrinsic_names", "") or "",
        detection_macro=d.get("detection_macro", "") or "",
        t_alias_in=d.get("t_alias_in"),
        v_var_in=d.get("v_var_in"),
        deprecated_in=d.get("deprecated_in"),
        notes=d.get("notes", "") or "",
        depends_on=tuple(d.get("depends_on") or ()),
        failure_reason=d.get("failure_reason"),
    )


# ---------------------------------------------------------------------------
# Data loader. data/*.json -> three lookups consumed by the builders:
#   HEADERS[header_name]         -> [Symbol, ...]
#   HEADER_SUBTITLES[header_name]-> long subtitle for title bar
#   HEADER_LEGEND_BLURBS[name]   -> short blurb for Legend sheet
# ---------------------------------------------------------------------------

def load_data(data_dir):
    """Loads per-header JSON files (data/<header>.json). Files whose
    basename starts with underscore (e.g. _roadmap.json) are skipped
    — those are non-header artifacts loaded separately."""
    headers = {}
    subtitles = {}
    legend_blurbs = {}
    for path in sorted(glob.glob(os.path.join(data_dir, "*.json"))):
        if os.path.basename(path).startswith("_"):
            continue
        with open(path) as f:
            obj = json.load(f)
        h = obj["header"]
        headers[h] = [symbol_from_dict(s) for s in obj["symbols"]]
        subtitles[h] = obj.get("subtitle", "")
        legend_blurbs[h] = obj.get("legend_blurb", "")
    return headers, subtitles, legend_blurbs


def load_roadmap(data_dir):
    """Loads data/_roadmap.json -> list of (priority, header, summary, blockers, est)."""
    path = os.path.join(data_dir, "_roadmap.json")
    with open(path) as f:
        items = json.load(f)
    return [
        (it["priority"], it["header"], it["summary"], it["blockers"], it["est_min_cpp"])
        for it in items
    ]


# ---------------------------------------------------------------------------
# Hardcoded: auxiliary-sheet metadata.
# These describe the workbook's structure, not its symbol data.
# Roadmap rows are loaded from data/_roadmap.json — see load_roadmap().
# ---------------------------------------------------------------------------

LEGEND_HEADER_ROW = ("Legend", "You are here. Colour key + column meanings.")

AUX_SHEET_BLURBS = [
    ("Pending Dependencies", "Symbols that need other restd modules before they reach full coverage. Auto-derived from the depends_on field in the per-header JSON files."),
    ("Coverage Failures",    "Detailed list of every symbol that does not reach C++98+ coverage, with the reason."),
    ("Roadmap",              "Priority-ordered plan for remaining work. The final entry is a REMINDER about deferred Master/Index sheets and in-tree stub status."),
]

LEGEND_COLUMN_MEANINGS = [
    ("Added to std",                 "C++ version when the symbol entered std::."),
    ("C++ version cells (7)",        "Support tier across C++98 → C++26."),
    ("constexpr cells (7)",          "Constexpr tier on the same axis: green = constexpr in std, yellow = constexpr in restd ahead of std, grey = exists but not constexpr, red = unavailable."),
    ("_t alias since",               "C++ version when std::<symbol>_t alias was added."),
    ("_v variable since",            "C++ version when std::<symbol>_v variable was added."),
    ("Deprecated in",                "C++ version when the symbol was deprecated; — when active."),
    ("Compiler intrinsic required?", "Yes when the symbol depends on a compiler builtin like __is_class, __is_enum, __underlying_type, alignof. Such rows show the intrinsic and a D_RESTD_HAS_* override macro. The implementation degrades safely (false_type, or symbol omission) when the intrinsic is absent."),
    ("Intrinsic(s)",                 "Compiler builtin(s) used by both std and restd; — when none."),
    ("restd detection macro",        "Predefinable macro that resolves to 1/0 based on intrinsic availability. Override before #include to force."),
    ("Notes / fallback behaviour",   "Behaviour when the trait is gated, intrinsic is missing, or the implementation has known limitations."),
]

LEGEND_LABEL_DESCRIPTIONS = {
    "std": "Symbol is part of std at this C++ version. Use std directly.",
    "restd": "restd back-ports the symbol on this C++ version (std does not yet have it).",
    "cx (green)": "In the constexpr block: symbol is constexpr in std at this version.",
    "cx (yellow)": "In the constexpr block: symbol is constexpr in restd ahead of std.",
    "— (red, support)": "Symbol unavailable on this C++ version.",
    "— (grey, constexpr)": "In the constexpr block: symbol exists at this tier but is not constexpr.",
    "Yes (intrinsic)": 'In "Compiler intrinsic required?": this symbol depends on a compiler builtin (no portable C++ implementation is possible). Each such row shows its intrinsic and the D_RESTD_HAS_* override macro.',
}


# ---------------------------------------------------------------------------
# Style palette (spreadsheet_style.md §5)
# ---------------------------------------------------------------------------

VERSIONS = ["C++98", "C++11", "C++14", "C++17", "C++20", "C++23", "C++26"]

FILL_STD = PatternFill("solid", fgColor="C6EFCE")
FILL_RESTD = PatternFill("solid", fgColor="FFEB9C")
FILL_UNAVAIL = PatternFill("solid", fgColor="FFC7CE")
FILL_NOT_CX = PatternFill("solid", fgColor="E7E6E6")
FILL_INTRINSIC = PatternFill("solid", fgColor="FCE4D6")
FILL_HEADER = PatternFill("solid", fgColor="305496")
FILL_ZEBRA = PatternFill("solid", fgColor="F8F9FA")
FILL_REMINDER = PatternFill("solid", fgColor="FFF2CC")

FONT_STD = Font(name="Calibri", size=10, color="006100")
FONT_RESTD = Font(name="Calibri", size=10, color="9C5700")
FONT_UNAVAIL = Font(name="Calibri", size=10, color="9C0006")
FONT_NOT_CX = Font(name="Calibri", size=10, color="595959", italic=True)
FONT_INTRINSIC = Font(name="Calibri", size=10, color="9C5700", bold=True)
FONT_HEADER = Font(name="Calibri", size=11, color="FFFFFF", bold=True)
FONT_TITLE = Font(name="Calibri", size=18, color="305496", bold=True)
FONT_SUBTITLE = Font(name="Calibri", size=10, color="595959", italic=True)
FONT_BODY = Font(name="Calibri", size=10)
FONT_REMINDER = Font(name="Calibri", size=10, color="7F6000", bold=True)
FONT_SECTION = Font(name="Calibri", size=11, color="305496", bold=True)

THIN_GREY = Side(style="thin", color="BFBFBF")
CELL_BORDER = Border(top=THIN_GREY, bottom=THIN_GREY, left=THIN_GREY, right=THIN_GREY)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# 25-column per-header layout (spreadsheet_style.md §4)
COLUMNS = [
    ("Std Header",                   12, "left"),
    ("Symbol",                       34, "left"),
    ("Category",                     26, "left"),
    ("Added to std",                 12, "center"),
    ("C++98",                         8, "center"),
    ("C++11",                         8, "center"),
    ("C++14",                         8, "center"),
    ("C++17",                         8, "center"),
    ("C++20",                         8, "center"),
    ("C++23",                         8, "center"),
    ("C++26",                         8, "center"),
    ("constexpr\nC++98",              9, "center"),
    ("constexpr\nC++11",              9, "center"),
    ("constexpr\nC++14",              9, "center"),
    ("constexpr\nC++17",              9, "center"),
    ("constexpr\nC++20",              9, "center"),
    ("constexpr\nC++23",              9, "center"),
    ("constexpr\nC++26",              9, "center"),
    ("_t alias since",               14, "center"),
    ("_v variable since",            14, "center"),
    ("Deprecated in",                14, "center"),
    ("Compiler intrinsic\nrequired?",13, "center"),
    ("Intrinsic(s)",                 34, "left"),
    ("restd detection macro",        38, "left"),
    ("Notes / fallback behaviour",   70, "left"),
]


# ---------------------------------------------------------------------------
# Cell-classification logic (spreadsheet_style.md §7)
# ---------------------------------------------------------------------------

def _ge(v_a, v_b):
    if v_a is None or v_b is None:
        return False
    return VERSIONS.index(v_a) >= VERSIONS.index(v_b)

def support_cell(version, std_in, restd_min):
    if std_in and _ge(version, std_in):
        return ("std", FILL_STD, FONT_STD)
    if restd_min and _ge(version, restd_min):
        return ("restd", FILL_RESTD, FONT_RESTD)
    return ("—", FILL_UNAVAIL, FONT_UNAVAIL)

def constexpr_cell(version, std_in, restd_min, cx_std, cx_restd):
    exists = (std_in and _ge(version, std_in)) or (restd_min and _ge(version, restd_min))
    if not exists:
        return ("—", FILL_UNAVAIL, FONT_UNAVAIL)
    if cx_std and _ge(version, cx_std):
        return ("cx", FILL_STD, FONT_STD)
    if cx_restd and _ge(version, cx_restd):
        return ("cx", FILL_RESTD, FONT_RESTD)
    return ("—", FILL_NOT_CX, FONT_NOT_CX)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def write_title_bar(ws, title, subtitle, col_count):
    end_col = get_column_letter(col_count)
    ws.merge_cells(f"A1:{end_col}1")
    ws["A1"] = title
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells(f"A2:{end_col}2")
    ws["A2"] = subtitle
    ws["A2"].font = FONT_SUBTITLE
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6


def write_header_row(ws, row_num, columns):
    for i, (label, width, _align) in enumerate(columns, 1):
        c = ws.cell(row=row_num, column=i, value=label)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row_num].height = 38


def build_data_sheet(wb, header_name, subtitle_override, symbols):
    ws = wb.create_sheet(header_name)
    write_title_bar(
        ws,
        header_name,
        f"Module surface: {len(symbols)} symbols. "
        "Green = std support. Yellow = restd back-port. Red = unavailable. "
        "The constexpr block tracks the same axis but for compile-time evaluation.",
        len(COLUMNS),
    )
    write_header_row(ws, 4, COLUMNS)

    for row_idx, sym in enumerate(symbols, start=5):
        body_zebra = (row_idx % 2 == 1)

        c = ws.cell(row=row_idx, column=1, value=header_name if row_idx == 5 else None)
        c.font = FONT_BODY; c.alignment = ALIGN_LEFT_TOP; c.border = CELL_BORDER
        if body_zebra: c.fill = FILL_ZEBRA

        c = ws.cell(row=row_idx, column=2, value=sym.name)
        c.font = FONT_BODY; c.alignment = ALIGN_LEFT_TOP; c.border = CELL_BORDER
        if body_zebra: c.fill = FILL_ZEBRA

        c = ws.cell(row=row_idx, column=3, value=sym.group)
        c.font = FONT_BODY; c.alignment = ALIGN_LEFT_TOP; c.border = CELL_BORDER
        if body_zebra: c.fill = FILL_ZEBRA

        c = ws.cell(row=row_idx, column=4, value=sym.std_in or "—")
        c.font = FONT_BODY; c.alignment = ALIGN_CENTER; c.border = CELL_BORDER
        if body_zebra: c.fill = FILL_ZEBRA

        for j, ver in enumerate(VERSIONS):
            col = 5 + j
            text, fill, font = support_cell(ver, sym.std_in, sym.restd_min)
            c = ws.cell(row=row_idx, column=col, value=text)
            c.font = font; c.fill = fill; c.alignment = ALIGN_CENTER; c.border = CELL_BORDER

        for j, ver in enumerate(VERSIONS):
            col = 12 + j
            text, fill, font = constexpr_cell(
                ver, sym.std_in, sym.restd_min,
                sym.constexpr_in_std, sym.constexpr_in_restd,
            )
            c = ws.cell(row=row_idx, column=col, value=text)
            c.font = font; c.fill = fill; c.alignment = ALIGN_CENTER; c.border = CELL_BORDER

        c = ws.cell(row=row_idx, column=19, value=sym.t_alias_in or "—")
        c.font = FONT_BODY; c.alignment = ALIGN_CENTER; c.border = CELL_BORDER
        if body_zebra: c.fill = FILL_ZEBRA

        c = ws.cell(row=row_idx, column=20, value=sym.v_var_in or "—")
        c.font = FONT_BODY; c.alignment = ALIGN_CENTER; c.border = CELL_BORDER
        if body_zebra: c.fill = FILL_ZEBRA

        c = ws.cell(row=row_idx, column=21, value=sym.deprecated_in or "—")
        c.font = FONT_BODY; c.alignment = ALIGN_CENTER; c.border = CELL_BORDER
        if body_zebra: c.fill = FILL_ZEBRA

        if sym.intrinsic_required:
            c = ws.cell(row=row_idx, column=22, value="Yes")
            c.font = FONT_INTRINSIC; c.fill = FILL_INTRINSIC
        else:
            c = ws.cell(row=row_idx, column=22, value="No")
            c.font = FONT_BODY
            if body_zebra: c.fill = FILL_ZEBRA
        c.alignment = ALIGN_CENTER; c.border = CELL_BORDER

        c = ws.cell(row=row_idx, column=23, value=sym.intrinsic_names or "—")
        c.font = FONT_BODY; c.alignment = ALIGN_LEFT_TOP; c.border = CELL_BORDER
        if body_zebra: c.fill = FILL_ZEBRA

        c = ws.cell(row=row_idx, column=24, value=sym.detection_macro or "—")
        c.font = FONT_BODY; c.alignment = ALIGN_LEFT_TOP; c.border = CELL_BORDER
        if body_zebra: c.fill = FILL_ZEBRA

        c = ws.cell(row=row_idx, column=25, value=sym.notes or "")
        c.font = FONT_BODY; c.alignment = ALIGN_LEFT_TOP; c.border = CELL_BORDER
        if body_zebra: c.fill = FILL_ZEBRA

        ws.row_dimensions[row_idx].height = 30

    ws.freeze_panes = "E5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(COLUMNS))}{4 + len(symbols)}"

    # NOTE: the JSON files carry a `subtitle` field (long per-header
    # blurb with phase history). The original v2 builder always used
    # the auto-generated "Module surface: N symbols..." subtitle and
    # ignored HEADER_SUBTITLES. Preserve that behavior — leave the
    # long subtitle in the JSON as metadata for now. If you want to
    # surface it on the data sheet, uncomment:
    # if subtitle_override:
    #     ws["A2"] = subtitle_override


def build_legend_sheet(wb, header_blurbs):
    """Builds the Legend sheet. header_blurbs: ordered list of (sheet_name, blurb)."""
    ws = wb.create_sheet("Legend")
    write_title_bar(
        ws,
        "restd module coverage matrix",
        "Single source of truth: data/*.json. Builder: build_workbook.py. Style guide: spreadsheet_style.md.",
        4,
    )

    def section(row, title):
        c = ws.cell(row=row, column=1, value=title)
        c.font = FONT_SECTION
        ws.row_dimensions[row].height = 22
        return row + 1

    def kv(row, label, descr):
        a = ws.cell(row=row, column=1, value=label)
        a.font = FONT_BODY; a.alignment = ALIGN_LEFT_TOP; a.border = CELL_BORDER
        b = ws.cell(row=row, column=2, value=descr)
        b.font = FONT_BODY; b.alignment = ALIGN_LEFT_TOP; b.border = CELL_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 28
        return row + 1

    def kv_swatch(row, label, sample_text, fill, font):
        a = ws.cell(row=row, column=1, value=label)
        a.font = FONT_BODY; a.alignment = ALIGN_LEFT_TOP; a.border = CELL_BORDER
        sw = ws.cell(row=row, column=2, value=sample_text)
        sw.fill = fill; sw.font = font; sw.alignment = ALIGN_CENTER; sw.border = CELL_BORDER
        b = ws.cell(row=row, column=3, value=LEGEND_LABEL_DESCRIPTIONS[label])
        b.font = FONT_BODY; b.alignment = ALIGN_LEFT_TOP; b.border = CELL_BORDER
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 28
        return row + 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 30

    row = 4
    row = section(row, "Sheets")
    # Legend self-reference, then per-header blurbs (alphabetical), then aux sheets.
    row = kv(row, *LEGEND_HEADER_ROW)
    for sheet_name, descr in header_blurbs:
        row = kv(row, sheet_name, descr)
    for sheet_name, descr in AUX_SHEET_BLURBS:
        row = kv(row, sheet_name, descr)

    row += 1
    row = section(row, "Colour key")
    row = kv_swatch(row, "std",                  "std", FILL_STD,    FONT_STD)
    row = kv_swatch(row, "restd",                "restd", FILL_RESTD,  FONT_RESTD)
    row = kv_swatch(row, "cx (green)",           "cx", FILL_STD,    FONT_STD)
    row = kv_swatch(row, "cx (yellow)",          "cx", FILL_RESTD,  FONT_RESTD)
    row = kv_swatch(row, "— (red, support)",     "—", FILL_UNAVAIL,FONT_UNAVAIL)
    row = kv_swatch(row, "— (grey, constexpr)",  "—", FILL_NOT_CX, FONT_NOT_CX)
    row = kv_swatch(row, "Yes (intrinsic)",      "Yes", FILL_INTRINSIC, FONT_INTRINSIC)

    row += 1
    row = section(row, "Other column meanings")
    for k, v in LEGEND_COLUMN_MEANINGS:
        row = kv(row, k, v)


def build_pending_deps(wb, headers):
    ws = wb.create_sheet("Pending Dependencies")
    write_title_bar(
        ws,
        "Pending Dependencies",
        "Symbols whose full coverage is blocked on a restd module that has not yet shipped. "
        "Auto-derived from the depends_on field in the per-header JSON files.",
        5,
    )
    columns = [
        ("Std Header", 14, "left"),
        ("Symbol", 36, "left"),
        ("Category", 26, "left"),
        ("Missing dependency", 36, "left"),
        ("Notes", 60, "left"),
    ]
    write_header_row(ws, 4, columns)
    row_idx = 5
    for header_name, syms in headers.items():
        for sym in syms:
            if sym.depends_on:
                missing = ", ".join(sym.depends_on)
                ws.cell(row=row_idx, column=1, value=header_name)
                ws.cell(row=row_idx, column=2, value=sym.name)
                ws.cell(row=row_idx, column=3, value=sym.group)
                ws.cell(row=row_idx, column=4, value=missing)
                ws.cell(row=row_idx, column=5, value=sym.notes or "")
                for col in range(1, 6):
                    c = ws.cell(row=row_idx, column=col)
                    c.font = FONT_BODY; c.alignment = ALIGN_LEFT_TOP; c.border = CELL_BORDER
                    if row_idx % 2 == 1: c.fill = FILL_ZEBRA
                ws.row_dimensions[row_idx].height = 30
                row_idx += 1
    ws.freeze_panes = "A5"
    if row_idx > 5:
        ws.auto_filter.ref = f"A4:E{row_idx - 1}"


def build_coverage_failures(wb, headers):
    ws = wb.create_sheet("Coverage Failures")
    write_title_bar(
        ws,
        "Coverage Failures",
        "Every symbol that does not reach full C++98+ coverage. "
        "Includes both partial coverage (restd_min > C++98) and unimplemented symbols (restd_min = none).",
        6,
    )
    columns = [
        ("Std Header", 14, "left"),
        ("Symbol", 36, "left"),
        ("Category", 26, "left"),
        ("Std introduced", 14, "center"),
        ("restd min", 16, "center"),
        ("Failure reason", 80, "left"),
    ]
    write_header_row(ws, 4, columns)
    row_idx = 5
    for header_name, syms in headers.items():
        for sym in syms:
            failed = (
                sym.failure_reason is not None
                or sym.restd_min is None
                or (sym.restd_min and sym.restd_min != "C++98")
            )
            if not failed:
                continue
            ws.cell(row=row_idx, column=1, value=header_name)
            ws.cell(row=row_idx, column=2, value=sym.name)
            ws.cell(row=row_idx, column=3, value=sym.group)
            ws.cell(row=row_idx, column=4, value=sym.std_in or "—")
            ws.cell(row=row_idx, column=5, value=sym.restd_min or "NOT IMPLEMENTED")
            ws.cell(row=row_idx, column=6, value=sym.failure_reason or "")
            for col in range(1, 7):
                c = ws.cell(row=row_idx, column=col)
                c.font = FONT_BODY
                c.alignment = ALIGN_CENTER if col in (4, 5) else ALIGN_LEFT_TOP
                c.border = CELL_BORDER
                if row_idx % 2 == 1: c.fill = FILL_ZEBRA
            ws.row_dimensions[row_idx].height = 30
            row_idx += 1
    ws.freeze_panes = "A5"
    if row_idx > 5:
        ws.auto_filter.ref = f"A4:F{row_idx - 1}"


def build_roadmap(wb, roadmap):
    ws = wb.create_sheet("Roadmap")
    write_title_bar(
        ws,
        "Roadmap",
        "Priority-ordered plan for remaining work. Lower priority number = sooner. "
        "The final row is a REMINDER about deferred Master and Index sheets.",
        5,
    )
    columns = [
        ("Priority", 10, "center"),
        ("Std Header", 14, "left"),
        ("Summary", 80, "left"),
        ("Blockers", 50, "left"),
        ("Est. min C++", 18, "center"),
    ]
    write_header_row(ws, 4, columns)

    for i, (priority, header, summary, blockers, est) in enumerate(roadmap):
        row_idx = 5 + i
        is_reminder = (header == "(meta)")
        ws.cell(row=row_idx, column=1, value=priority)
        ws.cell(row=row_idx, column=2, value=header)
        ws.cell(row=row_idx, column=3, value=summary)
        ws.cell(row=row_idx, column=4, value=blockers)
        ws.cell(row=row_idx, column=5, value=est)
        for col in range(1, 6):
            c = ws.cell(row=row_idx, column=col)
            if is_reminder:
                c.font = FONT_REMINDER
                c.fill = FILL_REMINDER
            else:
                c.font = FONT_BODY
                if row_idx % 2 == 1: c.fill = FILL_ZEBRA
            c.alignment = ALIGN_CENTER if col in (1, 5) else ALIGN_LEFT_TOP
            c.border = CELL_BORDER
        ws.row_dimensions[row_idx].height = 36 if is_reminder else 30
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:E{4 + len(roadmap)}"


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------

def build(data_dir):
    headers, subtitles, legend_blurbs = load_data(data_dir)
    roadmap = load_roadmap(data_dir)
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    # 1. Legend (uses ordered per-header blurbs)
    sorted_headers = sorted(headers.keys())
    header_blurbs = [(h, legend_blurbs.get(h, "")) for h in sorted_headers]
    build_legend_sheet(wb, header_blurbs)

    # 2. Per-header sheets, alphabetical (spreadsheet_style.md §3)
    for header in sorted_headers:
        build_data_sheet(wb, header, subtitles.get(header), headers[header])

    # 3. Pending Dependencies, Coverage Failures, Roadmap
    build_pending_deps(wb, headers)
    build_coverage_failures(wb, headers)
    build_roadmap(wb, roadmap)

    return wb


if __name__ == "__main__":
    wb = build(DATA_DIR)
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH}")
    print(f"  data:  {DATA_DIR}")
    print(f"  sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
