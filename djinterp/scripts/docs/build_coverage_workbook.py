#!/usr/bin/env python3
"""
Module coverage workbook builder (generic).

Reads JSON data describing the version-portability and constexpr-portability
of symbols in a code module, and writes an Excel workbook visualising
coverage across C++ language versions.

This is a project-agnostic generalisation of the original restd workbook
builder. The two universal axes are:
    - support  : at which C++ version does the symbol work?
    - constexpr: at which C++ version is it usable in constant evaluation?

Each axis may have up to two tiers (a "primary" baseline and a "secondary"
back-port). Restd uses both ("std" / "restd"). A project like djinterp
that is itself the library being measured uses only the primary tier; in
that case the secondary tier is disabled in _config.json and the workbook
renders single-coloured support cells (green = available, red = not).

Inputs (under --data-dir):
    _config.json        project-level configuration (required)
    _roadmap.json       roadmap entries for the To Do sheet (optional)
    *.json              one file per header / module unit (the data)

Output:
    --output            the .xlsx path

Usage:
    python3 build_coverage_workbook.py \\
        --data-dir docs/djinterp/functional/data \\
        --output   docs/djinterp/functional/functional_coverage.xlsx

The JSON layer remains the single source of truth. Everything visual lives
in the STYLE section near the top of this file plus the _config.json
file. Change either and the whole workbook re-themes.
"""

import argparse
import glob
import json
import os
from collections import namedtuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================================
# CLI
# ============================================================================

def parse_args():
    p = argparse.ArgumentParser(
        description="Build a coverage workbook from per-header JSON files.")
    p.add_argument("--data-dir", required=True,
                   help="Directory containing _config.json, _roadmap.json, "
                        "and one *.json per header.")
    p.add_argument("--output", required=True,
                   help="Path to write the .xlsx workbook to.")
    return p.parse_args()


# ============================================================================
# Config — defaults + loader
# ============================================================================

# Anything not overridden in _config.json falls back to these defaults.
# A project that wants the original restd look just leaves _config.json
# empty or omits it entirely.
DEFAULT_CONFIG = {
    # display
    "project_name":      "Project",
    "workbook_title":    "Module coverage matrix",
    "workbook_subtitle": "Per-header coverage of the project across C++ "
                         "versions, with constexpr eligibility tracked "
                         "separately.",
    # unit terminology ("header", "module", "translation unit", ...)
    # The singular noun appears in section labels (e.g. "Std Header" column,
    # "Per-header completeness summary"). Plural is used in title bars.
    "unit_label":        "Header",
    "unit_label_plural": "headers",

    # tiers: drives colour scheme and column labelling
    #   primary   : the baseline ("std", "djinterp", "boost", ...)
    #   secondary : optional back-port tier; null => single-tier project
    "tiers": {
        "primary":   {"label": "std"},
        "secondary": {"label": "restd"}
    },

    # How support cells are coloured. Two views of the same data:
    #   "baseline" (default): green = the PRIMARY baseline has it at this
    #              version, yellow = the SECONDARY back-port reaches it below
    #              the baseline, red = neither. Answers "where does the
    #              standard provide this?" — and is the only sensible mode for
    #              a single-tier project (no secondary), where green = present.
    #   "coverage": green = the SECONDARY tier (the library being measured,
    #              e.g. restd) covers it at this version, yellow = only the
    #              PRIMARY baseline has it here (a gap in our coverage), red =
    #              no coverage at all. Answers "where does OUR implementation
    #              cover this?". Only meaningful when a secondary tier exists;
    #              with no secondary it falls back to "baseline".
    "support_coloring": "baseline",

    # Category-column text casing:
    #   "preserve" (default): render group strings exactly as authored.
    #   "lower": lowercase each word EXCEPT all-caps tokens (acronyms like
    #            RAII, CTAD, SBO, ADL are left untouched).
    "category_case": "preserve",

    # which optional columns to render in each data sheet
    "columns": {
        "primary_in":       True,   # "Added to <primary>" — when std added it
        "support_matrix":   True,   # 7 columns "coverage" (always on)
        "constexpr_matrix": True,   # 7 columns "constexpr in" (always on)
        # merges constexpr signal into support cells: when a symbol is
        # supported AND constexpr at a given version, the cell text becomes
        # "cx" instead of the tier label. The constexpr_matrix block can
        # then be turned off to collapse the layout into a single 7-column
        # support block that doubles as the constexpr indicator.
        "support_shows_constexpr": False,
        "t_alias":          True,   # _t alias since
        "v_var":            True,   # _v variable since
        "deprecated":       True,   # deprecated in
        "intrinsic":        True,   # compiler intrinsic required?
        "intrinsic_names":  True,
        "detection_macro":  True,
        "notes":            True,
    },

    # C++ version axis (extend or trim per project as needed)
    "versions": ["C++98", "C++11", "C++14", "C++17", "C++20", "C++23", "C++26"],

    # which sheets are auto-emitted (set to false to suppress)
    "sheets": {
        "legend":         True,
        "todo":           True,
        "pending_deps":   True,
        "failures":       True,
        "reexports":      True,
    },
}


def load_config(data_dir, overrides=None):
    """Loads _config.json from data_dir, merging into DEFAULT_CONFIG. A
    missing _config.json is fine; defaults apply unchanged.

    If `overrides` (a dict) is given, it is deep-merged LAST — on top of both
    the defaults and any on-disk _config.json. This lets a caller bake a
    project config into code (e.g. restd.py) and skip the data-dir file
    entirely, while a stray _config.json can still contribute keys the
    overrides don't mention."""
    cfg = json.loads(json.dumps(DEFAULT_CONFIG))   # deep-copy
    path = os.path.join(data_dir, "_config.json")
    if os.path.isfile(path):
        with open(path) as f:
            user = json.load(f)
        _deep_merge(cfg, user)
    if overrides:
        _deep_merge(cfg, overrides)
    return cfg


def _deep_merge(target, src):
    """Recursive merge: src values override target; dicts are merged
    rather than replaced wholesale so a project can override just one
    sub-key (e.g. columns.notes) without restating the rest."""
    for k, v in src.items():
        if isinstance(v, dict) and isinstance(target.get(k), dict):
            _deep_merge(target[k], v)
        else:
            target[k] = v


# ============================================================================
# Symbol record — canonical schema with legacy aliases
# ============================================================================

# Canonical field names use "primary" / "secondary" so the schema is
# project-agnostic. The loader transparently accepts the original restd
# field names (std_in / restd_min / constexpr_in_std / constexpr_in_restd)
# so existing data files keep working unchanged.
Symbol = namedtuple("Symbol", [
    "name", "group",
    "primary_in", "secondary_min",
    "constexpr_in_primary", "constexpr_in_secondary",
    "intrinsic_required", "intrinsic_names", "detection_macro",
    "t_alias_in", "v_var_in", "deprecated_in",
    "notes", "depends_on", "failure_reason",
    "reexport",
])


def symbol_from_dict(d):
    def _pick(*keys):
        for k in keys:
            if k in d and d[k] is not None:
                return d[k]
        return None
    return Symbol(
        name                   = d["name"],
        group                  = d.get("group", "") or "",
        primary_in             = _pick("primary_in", "std_in"),
        secondary_min          = _pick("secondary_min", "restd_min"),
        constexpr_in_primary   = _pick("constexpr_in_primary", "constexpr_in_std"),
        constexpr_in_secondary = _pick("constexpr_in_secondary", "constexpr_in_restd"),
        intrinsic_required     = bool(d.get("intrinsic_required", False)),
        intrinsic_names        = d.get("intrinsic_names", "") or "",
        detection_macro        = d.get("detection_macro", "") or "",
        t_alias_in             = d.get("t_alias_in"),
        v_var_in               = d.get("v_var_in"),
        deprecated_in          = d.get("deprecated_in"),
        notes                  = d.get("notes", "") or "",
        depends_on             = tuple(d.get("depends_on") or ()),
        failure_reason         = d.get("failure_reason"),
        reexport               = bool(d.get("reexport", False)),
    )


# ============================================================================
# Data loader
# ============================================================================

def load_data(data_dir):
    headers       = {}
    subtitles     = {}
    legend_blurbs = {}
    for path in sorted(glob.glob(os.path.join(data_dir, "*.json"))):
        if os.path.basename(path).startswith("_"):
            continue
        with open(path) as f:
            obj = json.load(f)
        h = obj["header"]
        headers[h]       = [symbol_from_dict(s) for s in obj["symbols"]]
        subtitles[h]     = obj.get("subtitle", "")
        legend_blurbs[h] = obj.get("legend_blurb", "")
    return headers, subtitles, legend_blurbs


def load_roadmap(data_dir):
    path = os.path.join(data_dir, "_roadmap.json")
    if not os.path.isfile(path):
        return []
    with open(path) as f:
        return json.load(f)


# ============================================================================
# STYLE — every visual setting lives in this section.
# ============================================================================

COLORS = {
    # data-cell fills
    "primary":          "C6EFCE",   # green
    "secondary":        "FFEB9C",   # yellow
    "unavail":          "FFC7CE",   # red
    "not_cx":           "E7E6E6",   # grey
    "intrinsic":        "FCE4D6",   # peach

    # structural fills
    "header_fill":      "305496",
    "subheader_fill":   "8EA9DB",
    "section_fill":     "DDEBF7",
    "zebra":            "F8F9FA",
    "reminder_fill":    "FFF2CC",

    # To Do completeness gradient
    "todo_done":        "70AD47",
    "todo_high":        "C6EFCE",
    "todo_mid":         "FFEB9C",
    "todo_low":         "FFD580",
    "todo_started":     "F4A582",
    "todo_none":        "F8CBAD",

    # Blocked-symbol triage
    "blocked_ready":    "C6EFCE",
    "blocked_partial":  "FFEB9C",
    "blocked_full":     "FFC7CE",

    # text
    "text_primary":     "006100",
    "text_secondary":   "9C5700",
    "text_unavail":     "9C0006",
    "text_not_cx":      "595959",
    "text_intrinsic":   "9C5700",
    "text_header":      "FFFFFF",
    "text_title":       "305496",
    "text_subtitle":    "595959",
    "text_body":        "000000",
    "text_reminder":    "7F6000",
    "text_section":     "305496",
}

FONT_FAMILY = "Calibri"
FONT_SIZES  = {"title": 18, "section": 11, "header": 11,
               "subheader": 10, "body": 10, "subtitle": 10}

COLUMN_WIDTHS = {
    "std_header":      14,
    "symbol":          34,
    "category":        26,
    "added_in":        12,
    "version":          8,
    "t_alias":         14,
    "v_var":           14,
    "deprecated":      14,
    "intrinsic":       13,
    "intrinsic_names": 34,
    "detection_macro": 38,
    "notes":           70,
}

ROW_HEIGHTS = {"title": 26, "subtitle": 18, "spacer": 6,
               "group_hdr": 22, "leaf_hdr": 28, "data": 30,
               "section": 22, "reminder": 36}

BORDER_THIN_COLOR  = "BFBFBF"
BORDER_THICK_COLOR = "305496"
BORDER_THICK_STYLE = "medium"

SHEET_ORDER_FIRST = ["Legend", "To Do"]
SHEET_ORDER_LAST  = ["Pending Dependencies", "Coverage Failures"]


# ============================================================================
# Derived openpyxl-typed style objects
# ============================================================================

def _fill(c):  return PatternFill("solid", fgColor=COLORS[c])

FILLS = {k: _fill(k) for k in (
    "primary", "secondary", "unavail", "not_cx", "intrinsic",
    "header_fill", "subheader_fill", "section_fill", "zebra", "reminder_fill",
    "todo_done", "todo_high", "todo_mid", "todo_low", "todo_started", "todo_none",
    "blocked_ready", "blocked_partial", "blocked_full",
)}

def _font(color_key, size=None, bold=False, italic=False):
    return Font(name=FONT_FAMILY,
                size=size or FONT_SIZES["body"],
                color=COLORS[color_key], bold=bold, italic=italic)

FONTS = {
    "primary":   _font("text_primary"),
    "secondary": _font("text_secondary"),
    "unavail":   _font("text_unavail"),
    "not_cx":    _font("text_not_cx", italic=True),
    "intrinsic": _font("text_intrinsic", bold=True),
    "header":    _font("text_header",   size=FONT_SIZES["header"],    bold=True),
    "subheader": _font("text_header",   size=FONT_SIZES["subheader"], bold=True),
    "title":     _font("text_title",    size=FONT_SIZES["title"],     bold=True),
    "subtitle":  _font("text_subtitle", size=FONT_SIZES["subtitle"],  italic=True),
    "body":      _font("text_body"),
    "reminder":  _font("text_reminder", bold=True),
    "section":   _font("text_section",  size=FONT_SIZES["section"],   bold=True),
}

SIDE_THIN  = Side(style="thin",             color=BORDER_THIN_COLOR)
SIDE_THICK = Side(style=BORDER_THICK_STYLE, color=BORDER_THICK_COLOR)
BORDER_ALL_THIN = Border(top=SIDE_THIN, bottom=SIDE_THIN,
                         left=SIDE_THIN, right=SIDE_THIN)

ALIGN_CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_LEFT_TOP = Alignment(horizontal="left",   vertical="top",    wrap_text=True)


# ============================================================================
# Cell classification (config-aware)
# ============================================================================

def _ge(versions, v_a, v_b):
    if v_a is None or v_b is None:
        return False
    if v_a not in versions or v_b not in versions:
        return False
    return versions.index(v_a) >= versions.index(v_b)


def display_category(cfg, group):
    """Renders a category/group string per cfg.category_case.

    In "lower" mode each whitespace-separated token is lowercased unless it
    looks like an acronym — an all-caps token of >=2 chars with at least two
    letters and no lowercase (RAII, CTAD, SBO, ADL, PMF, I/O) — which is left
    untouched. "preserve" (default) returns the string unchanged.
    """
    if not group or cfg.get("category_case", "preserve") != "lower":
        return group
    out = []
    for tok in group.split(" "):
        letters = [ch for ch in tok if ch.isalpha()]
        is_acronym = (len(tok) >= 2 and len(letters) >= 2
                      and tok == tok.upper() and tok != tok.lower())
        out.append(tok if is_acronym else tok.lower())
    return " ".join(out)


def support_cell(cfg, version, primary_in, secondary_min,
                 cx_primary=None, cx_secondary=None):
    """Returns (text, fill, font) for a single support-axis cell.

    Two colouring views, selected by cfg.support_coloring:

      "baseline" (default): green = primary baseline has it here, yellow =
          secondary back-port reaches it below the baseline, red = neither.

      "coverage": green = the secondary tier (the measured library) covers
          it here, yellow = only the primary baseline has it (a coverage
          gap), red = no coverage. Falls back to "baseline" if there is no
          secondary tier.

    When cfg.columns.support_shows_constexpr is True the cell text becomes
    "cx" wherever the symbol is constexpr at this version (per whichever
    tier owns the cell), and "" otherwise — so a single coloured block
    doubles as the constexpr indicator. The fill always carries the support
    signal.
    """
    versions = cfg["versions"]
    has_secondary = cfg["tiers"].get("secondary") is not None
    annotate_cx  = cfg["columns"].get("support_shows_constexpr", False)
    coverage_mode = (has_secondary
                     and cfg.get("support_coloring", "baseline") == "coverage")
    primary_supported   = primary_in    and _ge(versions, version, primary_in)
    secondary_supported = (has_secondary and secondary_min and
                           _ge(versions, version, secondary_min))

    def _cell(owner_constexpr_in, fill_key, font_key, label):
        # fill_key/font_key index FILLS/FONTS by colour ("primary" == green,
        # "secondary" == yellow); the owner_constexpr_in drives the "cx" text.
        if annotate_cx:
            is_cx = owner_constexpr_in and _ge(versions, version,
                                               owner_constexpr_in)
            text = "cx" if is_cx else ""
        else:
            text = label
        return (text, FILLS[fill_key], FONTS[font_key])

    if coverage_mode:
        # Library-centric: secondary coverage is the headline (green); a
        # version that only the baseline reaches is the lesser state (yellow).
        if secondary_supported:
            return _cell(cx_secondary, "primary", "primary",
                         cfg["tiers"]["secondary"]["label"])
        if primary_supported:
            return _cell(cx_primary, "secondary", "secondary",
                         cfg["tiers"]["primary"]["label"])
        return ("\u2014", FILLS["unavail"], FONTS["unavail"])

    # Baseline-centric (original): primary green, secondary yellow.
    if primary_supported:
        return _cell(cx_primary, "primary", "primary",
                     cfg["tiers"]["primary"]["label"])
    if secondary_supported:
        return _cell(cx_secondary, "secondary", "secondary",
                     cfg["tiers"]["secondary"]["label"])
    return ("\u2014", FILLS["unavail"], FONTS["unavail"])


def constexpr_cell(cfg, version, primary_in, secondary_min,
                   cx_primary, cx_secondary):
    """Returns (text, fill, font) for a single constexpr-axis cell."""
    versions = cfg["versions"]
    has_secondary = cfg["tiers"].get("secondary") is not None
    exists = (primary_in and _ge(versions, version, primary_in)) or \
             (has_secondary and secondary_min and
              _ge(versions, version, secondary_min))
    if not exists:
        return ("\u2014", FILLS["unavail"], FONTS["unavail"])
    if cx_primary and _ge(versions, version, cx_primary):
        return ("cx", FILLS["primary"], FONTS["primary"])
    if has_secondary and cx_secondary and _ge(versions, version, cx_secondary):
        return ("cx", FILLS["secondary"], FONTS["secondary"])
    return ("\u2014", FILLS["not_cx"], FONTS["not_cx"])


# ============================================================================
# Dynamic column layout (driven by config)
# ============================================================================
# COLUMN_LEAVES / COLUMN_GROUPS were constants in the original. They are now
# computed per-project from cfg["columns"]. Each entry is:
#   (slot_key, leaf_label, width_key, leaf_alignment)
# Group separators (thick borders) are computed alongside.

def build_layout(cfg):
    """Returns (leaves, groups, separators_after, slot_to_col, col_to_slot,
    thick_right_cols, n_cols)."""
    cols = cfg["columns"]
    versions = cfg["versions"]
    has_secondary = cfg["tiers"].get("secondary") is not None
    primary_label = cfg["tiers"]["primary"]["label"]

    leaves = []
    groups = []
    seps   = set()

    # Always-on: Header column, Symbol column, Category column.
    leaves.append(("std_header", cfg["unit_label"],  "std_header", "left"))
    groups.append(("std_header", "std_header", None))
    leaves.append(("symbol",     "Symbol",           "symbol",     "left"))
    groups.append(("symbol",     "symbol",     None))
    leaves.append(("category",   "Category",         "category",   "left"))
    groups.append(("category",   "category",   None))
    # Thick separator between the name columns and the data matrix that
    # follows. This is visual punctuation: everything from here on is
    # version-axis data; everything before is row identity.
    seps.add("category")

    # Optional: "Added to <primary>" (when the symbol entered the primary tier)
    if cols.get("primary_in", True):
        leaves.append(("added_in", primary_label, "added_in", "center"))
        groups.append(("added_in", "added_in", f"Added to"))
        seps.add("added_in")

    # Support matrix — 7 (or len(versions)) columns
    if cols.get("support_matrix", True):
        for v in versions:
            slot = f"cov_{_slug(v)}"
            leaves.append((slot, v, "version", "left"))
        groups.append((f"cov_{_slug(versions[0])}",
                       f"cov_{_slug(versions[-1])}",
                       "coverage"))
        seps.add(f"cov_{_slug(versions[-1])}")

    # Constexpr matrix
    if cols.get("constexpr_matrix", True):
        for v in versions:
            slot = f"cx_{_slug(v)}"
            leaves.append((slot, v, "version", "left"))
        groups.append((f"cx_{_slug(versions[0])}",
                       f"cx_{_slug(versions[-1])}",
                       "constexpr in"))
        seps.add(f"cx_{_slug(versions[-1])}")

    # Std-library-specific optional columns
    if cols.get("t_alias",     False):
        leaves.append(("t_alias",    "since", "t_alias",    "center"))
        groups.append(("t_alias",    "t_alias",    "_t alias"))
    if cols.get("v_var",       False):
        leaves.append(("v_var",      "since", "v_var",      "center"))
        groups.append(("v_var",      "v_var",      "_v variable"))
    if cols.get("deprecated",  False):
        leaves.append(("deprecated", "in",    "deprecated", "center"))
        groups.append(("deprecated", "deprecated", "Deprecated"))
    if cols.get("intrinsic",   False):
        leaves.append(("intrinsic", "required?", "intrinsic", "center"))
        groups.append(("intrinsic", "intrinsic", "Compiler intrinsic"))
    if cols.get("intrinsic_names", False):
        leaves.append(("intrinsic_names", "Intrinsic(s)", "intrinsic_names", "left"))
        groups.append(("intrinsic_names", "intrinsic_names", None))
    if cols.get("detection_macro", False):
        leaves.append(("detection_macro", "Detection macro",
                       "detection_macro", "left"))
        groups.append(("detection_macro", "detection_macro", None))
    if cols.get("notes", True):
        leaves.append(("notes", "Notes / fallback behaviour", "notes", "left"))
        groups.append(("notes", "notes", None))

    slot_to_col = {leaf[0]: i + 1 for i, leaf in enumerate(leaves)}
    col_to_slot = {i + 1: leaf[0] for i, leaf in enumerate(leaves)}
    thick_right = {slot_to_col[s] for s in seps}
    # For every column with a thick right border, also paint the next
    # column's left border thick. This guarantees the separator renders
    # consistently regardless of which cell's fill / border is drawn first
    # — openpyxl + Excel can otherwise produce asymmetric edges where one
    # side is thick and the other thin, depending on cell fills.
    thick_left  = {col + 1 for col in thick_right
                   if col + 1 <= len(leaves)}

    return {
        "leaves":      leaves,
        "groups":      groups,
        "seps":        seps,
        "slot_to_col": slot_to_col,
        "col_to_slot": col_to_slot,
        "thick_right": thick_right,
        "thick_left":  thick_left,
        "n_cols":      len(leaves),
    }


def _slug(version):
    """C++14 -> c14, C++98 -> c98 — used to build slot keys for version cols."""
    return version.lower().replace("c++", "c")


def data_cell_border(layout, col_idx):
    right = SIDE_THICK if col_idx in layout["thick_right"] else SIDE_THIN
    left  = SIDE_THICK if col_idx in layout["thick_left"]  else SIDE_THIN
    return Border(top=SIDE_THIN, bottom=SIDE_THIN, left=left, right=right)


# ============================================================================
# Title bar + header block
# ============================================================================

def write_title_bar(ws, title, subtitle, col_count):
    end_col = get_column_letter(col_count)
    ws.merge_cells(f"A1:{end_col}1")
    ws["A1"] = title
    ws["A1"].font      = FONTS["title"]
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = ROW_HEIGHTS["title"]

    ws.merge_cells(f"A2:{end_col}2")
    ws["A2"] = subtitle
    ws["A2"].font      = FONTS["subtitle"]
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = ROW_HEIGHTS["subtitle"]
    ws.row_dimensions[3].height = ROW_HEIGHTS["spacer"]


def _write_header_block(ws, cfg, layout):
    leaves      = layout["leaves"]
    groups      = layout["groups"]
    slot_to_col = layout["slot_to_col"]
    col_to_slot = layout["col_to_slot"]

    # Column widths
    for slot, _leaf, width_key, _align in leaves:
        col_letter = get_column_letter(slot_to_col[slot])
        ws.column_dimensions[col_letter].width = COLUMN_WIDTHS[width_key]

    # Row 4 — group headers
    for first_slot, last_slot, label in groups:
        first = slot_to_col[first_slot]
        last  = slot_to_col[last_slot]
        first_letter = get_column_letter(first)
        last_letter  = get_column_letter(last)

        if label is None:
            # Single-column slot, merge row 4 with row 5
            ws.merge_cells(f"{first_letter}4:{last_letter}5")
            leaf_label = next(l[1] for l in leaves if l[0] == first_slot)
            c = ws.cell(row=4, column=first, value=leaf_label)
            c.font = FONTS["header"]; c.fill = FILLS["header_fill"]
            c.alignment = ALIGN_CENTER
        else:
            if first != last:
                ws.merge_cells(start_row=4, end_row=4,
                               start_column=first, end_column=last)
            c = ws.cell(row=4, column=first, value=label)
            c.font = FONTS["subheader"]; c.fill = FILLS["subheader_fill"]
            c.alignment = ALIGN_CENTER

    # Row 5 — leaf headers (only for groups that have a row-4 label).
    for first_slot, last_slot, label in groups:
        if label is None:
            continue
        first = slot_to_col[first_slot]
        last  = slot_to_col[last_slot]
        for col in range(first, last + 1):
            slot_at_col = col_to_slot[col]
            leaf_label  = next(l[1] for l in leaves if l[0] == slot_at_col)
            c = ws.cell(row=5, column=col, value=leaf_label)
            c.font = FONTS["header"]; c.fill = FILLS["header_fill"]
            slot_align = next(l[3] for l in leaves if l[0] == slot_at_col)
            c.alignment = ALIGN_LEFT if slot_align == "left" else ALIGN_CENTER

    # Borders on both header rows
    for row in (4, 5):
        for col in range(1, len(leaves) + 1):
            ws.cell(row=row, column=col).border = data_cell_border(layout, col)

    ws.row_dimensions[4].height = ROW_HEIGHTS["group_hdr"]
    ws.row_dimensions[5].height = ROW_HEIGHTS["leaf_hdr"]


# ============================================================================
# Per-header data sheet
# ============================================================================

def build_data_sheet(wb, cfg, layout, header_name, subtitle, symbols):
    ws = wb.create_sheet(_sheet_safe(header_name))
    n_cols = layout["n_cols"]

    has_secondary = cfg["tiers"].get("secondary") is not None
    coverage_mode = (has_secondary
                     and cfg.get("support_coloring", "baseline") == "coverage")
    if coverage_mode:
        default_subtitle = (
            f"Module surface: {len(symbols)} symbols. "
            f"Green = {cfg['tiers']['secondary']['label']} coverage. "
            f"Yellow = {cfg['tiers']['primary']['label']} only "
            f"(coverage gap). Red = no coverage. ")
    else:
        default_subtitle = (
            f"Module surface: {len(symbols)} symbols. "
            f"Green = {cfg['tiers']['primary']['label']} support. ")
        if has_secondary:
            default_subtitle += (
                f"Yellow = {cfg['tiers']['secondary']['label']} back-port. ")
        default_subtitle += "Red = unavailable. "
    if cfg["columns"].get("constexpr_matrix", True):
        default_subtitle += ("The constexpr block tracks the same axis but "
                             "for constant evaluation.")
    elif cfg["columns"].get("support_shows_constexpr", False):
        default_subtitle += ("Cells containing \"cx\" mark versions where "
                             "the symbol is usable in a constant expression.")

    write_title_bar(ws, header_name, subtitle or default_subtitle, n_cols)
    _write_header_block(ws, cfg, layout)

    DATA_START = 6
    slot_to_col = layout["slot_to_col"]
    versions = cfg["versions"]

    for row_idx, sym in enumerate(symbols, start=DATA_START):
        zebra = (row_idx % 2 == 1)

        def _set(col, value, font=FONTS["body"], align=ALIGN_LEFT_TOP,
                 fill_zebra=True):
            c = ws.cell(row=row_idx, column=col, value=value)
            c.font      = font
            c.alignment = align
            c.border    = data_cell_border(layout, col)
            if fill_zebra and zebra:
                c.fill = FILLS["zebra"]
            return c

        # Always-on columns
        _set(slot_to_col["std_header"],
             header_name if row_idx == DATA_START else None)
        _set(slot_to_col["symbol"],   sym.name)
        _set(slot_to_col["category"], display_category(cfg, sym.group))

        if "added_in" in slot_to_col:
            _set(slot_to_col["added_in"],
                 sym.primary_in or "\u2014", align=ALIGN_CENTER)

        # Support matrix
        if cfg["columns"].get("support_matrix", True):
            base = slot_to_col[f"cov_{_slug(versions[0])}"]
            for j, ver in enumerate(versions):
                col = base + j
                text, fill, font = support_cell(
                    cfg, ver, sym.primary_in, sym.secondary_min,
                    sym.constexpr_in_primary, sym.constexpr_in_secondary)
                c = ws.cell(row=row_idx, column=col, value=text)
                c.font = font; c.fill = fill; c.alignment = ALIGN_CENTER
                c.border = data_cell_border(layout, col)

        # Constexpr matrix
        if cfg["columns"].get("constexpr_matrix", True):
            base = slot_to_col[f"cx_{_slug(versions[0])}"]
            for j, ver in enumerate(versions):
                col = base + j
                text, fill, font = constexpr_cell(
                    cfg, ver, sym.primary_in, sym.secondary_min,
                    sym.constexpr_in_primary, sym.constexpr_in_secondary)
                c = ws.cell(row=row_idx, column=col, value=text)
                c.font = font; c.fill = fill; c.alignment = ALIGN_CENTER
                c.border = data_cell_border(layout, col)

        # Optional metadata columns
        if "t_alias" in slot_to_col:
            _set(slot_to_col["t_alias"],
                 sym.t_alias_in or "\u2014", align=ALIGN_CENTER)
        if "v_var" in slot_to_col:
            _set(slot_to_col["v_var"],
                 sym.v_var_in or "\u2014", align=ALIGN_CENTER)
        if "deprecated" in slot_to_col:
            _set(slot_to_col["deprecated"],
                 sym.deprecated_in or "\u2014", align=ALIGN_CENTER)
        if "intrinsic" in slot_to_col:
            if sym.intrinsic_required:
                c = ws.cell(row=row_idx,
                            column=slot_to_col["intrinsic"], value="Yes")
                c.font = FONTS["intrinsic"]; c.fill = FILLS["intrinsic"]
                c.alignment = ALIGN_CENTER
                c.border = data_cell_border(layout, slot_to_col["intrinsic"])
            else:
                _set(slot_to_col["intrinsic"], "No", align=ALIGN_CENTER)
        if "intrinsic_names" in slot_to_col:
            _set(slot_to_col["intrinsic_names"],
                 sym.intrinsic_names or "\u2014")
        if "detection_macro" in slot_to_col:
            _set(slot_to_col["detection_macro"],
                 sym.detection_macro or "\u2014")
        if "notes" in slot_to_col:
            _set(slot_to_col["notes"], sym.notes or "")

        ws.row_dimensions[row_idx].height = ROW_HEIGHTS["data"]

    # Freeze: rows above + first three name columns
    freeze_col_letter = get_column_letter(slot_to_col.get(
        f"cov_{_slug(versions[0])}", 4))
    ws.freeze_panes = f"{freeze_col_letter}6"
    ws.auto_filter.ref = (
        f"A5:{get_column_letter(n_cols)}{DATA_START - 1 + len(symbols)}")

    # Per-sheet failure section
    failures = [s for s in symbols if _is_failure(cfg, s)]
    if failures:
        _write_local_failures(ws, cfg, layout, header_name, failures,
                              DATA_START + len(symbols))


def _is_failure(cfg, sym):
    """A 'failure' is a symbol whose secondary_min isn't the floor version,
    OR has an explicit failure_reason, OR isn't implemented at all.

    For single-tier projects (no secondary tier), this collapses to: doesn't
    reach the floor on the primary tier, OR has a failure_reason."""
    floor = cfg["versions"][0]
    has_secondary = cfg["tiers"].get("secondary") is not None
    if sym.failure_reason is not None:
        return True
    if has_secondary:
        return sym.secondary_min is None or sym.secondary_min != floor
    return sym.primary_in is None or sym.primary_in != floor


def _sheet_safe(name):
    """Excel sheet names cannot contain: : \\ / ? * [ ] and must be <= 31
    chars. Filenames like 'extractor.hpp' are fine; '<algorithm>' becomes
    'algorithm' for readability."""
    n = name
    for ch in (":", "\\", "/", "?", "*", "[", "]"):
        n = n.replace(ch, "_")
    if n.startswith("<") and n.endswith(">"):
        n = n[1:-1]
    return n[:31]


def _write_local_failures(ws, cfg, layout, header_name, failures, start_row):
    """Mini failure table appended below the main data table on a sheet."""
    row = start_row + 2
    n_cols = layout["n_cols"]
    slot_to_col = layout["slot_to_col"]
    floor = cfg["versions"][0]

    banner_end = get_column_letter(n_cols)
    ws.merge_cells(f"A{row}:{banner_end}{row}")
    c = ws.cell(row=row, column=1,
                value=f"Coverage Failures (this sheet only) — "
                      f"{len(failures)} symbol(s) below 100% {floor} coverage")
    c.font = FONTS["section"]; c.fill = FILLS["section_fill"]
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[row].height = ROW_HEIGHTS["section"]
    row += 1

    # Heading layout follows the main columns so widths align nicely.
    sym_col   = slot_to_col["symbol"]
    cat_col   = slot_to_col["category"]
    added_col = slot_to_col.get("added_in", cat_col + 1)
    cov_col   = slot_to_col.get(f"cov_{_slug(cfg['versions'][0])}", added_col + 1)

    headings = ["Symbol", "Category", "Added to", "Floor reached", "Failure reason"]
    starts   = [1, sym_col + 1, cat_col + 1, added_col + 1, cov_col + 1]
    ends     = [sym_col, cat_col, added_col, cov_col, n_cols]

    for label, s, e in zip(headings, starts, ends):
        if s != e:
            ws.merge_cells(start_row=row, end_row=row,
                           start_column=s, end_column=e)
        c = ws.cell(row=row, column=s, value=label)
        c.font = FONTS["header"]; c.fill = FILLS["header_fill"]
        c.alignment = ALIGN_CENTER; c.border = BORDER_ALL_THIN
    ws.row_dimensions[row].height = ROW_HEIGHTS["leaf_hdr"]
    row += 1

    for i, sym in enumerate(failures):
        zebra = (i % 2 == 0)
        floor_reached = sym.secondary_min if cfg["tiers"].get("secondary") \
                        else sym.primary_in
        cells_data = [
            (sym.name,                    starts[0], ends[0], ALIGN_LEFT_TOP),
            (display_category(cfg, sym.group), starts[1], ends[1], ALIGN_LEFT_TOP),
            (sym.primary_in or "\u2014",  starts[2], ends[2], ALIGN_CENTER),
            (floor_reached or "NOT IMPLEMENTED",
                                          starts[3], ends[3], ALIGN_CENTER),
            (sym.failure_reason or "",    starts[4], ends[4], ALIGN_LEFT_TOP),
        ]
        for value, s, e, align in cells_data:
            if s != e:
                ws.merge_cells(start_row=row, end_row=row,
                               start_column=s, end_column=e)
            c = ws.cell(row=row, column=s, value=value)
            c.font = FONTS["body"]; c.alignment = align
            c.border = BORDER_ALL_THIN
            if zebra: c.fill = FILLS["zebra"]
        ws.row_dimensions[row].height = ROW_HEIGHTS["data"]
        row += 1


# ============================================================================
# Legend sheet
# ============================================================================

def build_legend_sheet(wb, cfg, header_blurbs):
    ws = wb.create_sheet("Legend")
    primary_label   = cfg["tiers"]["primary"]["label"]
    secondary_label = (cfg["tiers"]["secondary"]["label"]
                       if cfg["tiers"].get("secondary") else None)

    write_title_bar(
        ws,
        cfg["workbook_title"],
        f"Single source of truth: data/*.json. Builder: this script. "
        f"All visual settings live in the STYLE section of the builder, "
        f"plus _config.json for project-level terminology.",
        4)

    def section(row, title):
        c = ws.cell(row=row, column=1, value=title)
        c.font = FONTS["section"]
        ws.row_dimensions[row].height = ROW_HEIGHTS["section"]
        return row + 1

    def kv(row, label, descr):
        a = ws.cell(row=row, column=1, value=label)
        a.font = FONTS["body"]; a.alignment = ALIGN_LEFT_TOP
        a.border = BORDER_ALL_THIN
        b = ws.cell(row=row, column=2, value=descr)
        b.font = FONTS["body"]; b.alignment = ALIGN_LEFT_TOP
        b.border = BORDER_ALL_THIN
        ws.merge_cells(start_row=row, start_column=2,
                       end_row=row, end_column=4)
        ws.row_dimensions[row].height = 28
        return row + 1

    def kv_swatch(row, label, sample_text, fill, font, descr):
        a = ws.cell(row=row, column=1, value=label)
        a.font = FONTS["body"]; a.alignment = ALIGN_LEFT_TOP
        a.border = BORDER_ALL_THIN
        sw = ws.cell(row=row, column=2, value=sample_text)
        sw.fill = fill; sw.font = font; sw.alignment = ALIGN_CENTER
        sw.border = BORDER_ALL_THIN
        b = ws.cell(row=row, column=3, value=descr)
        b.font = FONTS["body"]; b.alignment = ALIGN_LEFT_TOP
        b.border = BORDER_ALL_THIN
        ws.merge_cells(start_row=row, start_column=3,
                       end_row=row, end_column=4)
        ws.row_dimensions[row].height = 28
        return row + 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 30

    row = 4
    row = section(row, "Sheets")
    row = kv(row, "Legend",
             "You are here. Colour key + column meanings.")
    for sheet_name, descr in header_blurbs:
        row = kv(row, _sheet_safe(sheet_name), descr)
    if cfg["sheets"].get("todo"):
        row = kv(row, "To Do",
                 "Per-header completeness summary + dependency triage.")
    if cfg["sheets"].get("pending_deps"):
        row = kv(row, "Pending Dependencies",
                 "Symbols blocked on other not-yet-shipped modules. "
                 "Auto-derived from depends_on.")
    if cfg["sheets"].get("failures"):
        row = kv(row, "Coverage Failures",
                 "Every symbol that does not reach the language floor, with "
                 "the reason.")

    row += 1
    row = section(row, "Colour key")
    annotate_cx = cfg["columns"].get("support_shows_constexpr", False)
    coverage_mode = (bool(secondary_label)
                     and cfg.get("support_coloring", "baseline") == "coverage")

    if coverage_mode:
        # Library-centric colours: green = secondary (e.g. restd) covers it,
        # yellow = only the primary baseline (std) has it, red = no coverage.
        # "cx" marks constant-evaluation coverage when the support/constexpr
        # blocks are merged.
        if annotate_cx:
            row = kv_swatch(row, "(blank)", "",
                            FILLS["primary"], FONTS["primary"],
                            f"{secondary_label} covers the symbol at this "
                            f"version (not usable in a constant expression).")
            row = kv_swatch(row, "cx", "cx",
                            FILLS["primary"], FONTS["primary"],
                            f"{secondary_label} covers it AND it is constexpr "
                            f"at this version.")
            row = kv_swatch(row, f"(blank, {primary_label} only)", "",
                            FILLS["secondary"], FONTS["secondary"],
                            f"Only {primary_label} has it at this version — a "
                            f"gap in {secondary_label} coverage (not constexpr).")
            row = kv_swatch(row, f"cx ({primary_label} only)", "cx",
                            FILLS["secondary"], FONTS["secondary"],
                            f"Only {primary_label} has it, but it is constexpr "
                            f"there.")
        else:
            row = kv_swatch(row, secondary_label, secondary_label,
                            FILLS["primary"], FONTS["primary"],
                            f"{secondary_label} covers the symbol at this "
                            f"version.")
            row = kv_swatch(row, f"{primary_label} only", primary_label,
                            FILLS["secondary"], FONTS["secondary"],
                            f"Only {primary_label} has it here — a gap in "
                            f"{secondary_label} coverage.")
        row = kv_swatch(row, "-- (red)", "\u2014",
                        FILLS["unavail"], FONTS["unavail"],
                        "No coverage on this C++ version.")
    elif annotate_cx:
        # Merged-cell layout, baseline colouring: support and constexpr share
        # one block, so the legend describes the cell states rather than the
        # support/constexpr pair.
        row = kv_swatch(row, "(blank)", "",
                        FILLS["primary"], FONTS["primary"],
                        f"Symbol is supported in {primary_label} at this "
                        f"version but is NOT constexpr. Green fill, no text.")
        row = kv_swatch(row, "cx", "cx",
                        FILLS["primary"], FONTS["primary"],
                        f"Symbol is supported AND constexpr in "
                        f"{primary_label} at this version.")
        if secondary_label:
            row = kv_swatch(row, f"(blank, {secondary_label})", "",
                            FILLS["secondary"], FONTS["secondary"],
                            f"{secondary_label} back-ports the symbol at "
                            f"this version (not constexpr).")
            row = kv_swatch(row, f"cx ({secondary_label})", "cx",
                            FILLS["secondary"], FONTS["secondary"],
                            f"{secondary_label} back-port AND constexpr at "
                            f"this version.")
        row = kv_swatch(row, "-- (red)", "\u2014",
                        FILLS["unavail"], FONTS["unavail"],
                        "Symbol unavailable on this C++ version.")
    else:
        # Original two-block layout: support and constexpr are separate
        # matrices, so the legend describes them as paired key/value entries.
        row = kv_swatch(row, primary_label, primary_label,
                        FILLS["primary"], FONTS["primary"],
                        f"Symbol is supported natively in {primary_label} at "
                        f"this version.")
        if secondary_label:
            row = kv_swatch(row, secondary_label, secondary_label,
                            FILLS["secondary"], FONTS["secondary"],
                            f"{secondary_label} back-ports the symbol at this "
                            f"version ({primary_label} does not yet have it).")
        row = kv_swatch(row, f"cx ({primary_label})", "cx",
                        FILLS["primary"], FONTS["primary"],
                        f"In the constexpr block: symbol is constexpr in "
                        f"{primary_label} at this version.")
        if secondary_label:
            row = kv_swatch(row, f"cx ({secondary_label})", "cx",
                            FILLS["secondary"], FONTS["secondary"],
                            f"In the constexpr block: symbol is constexpr in "
                            f"{secondary_label} ahead of {primary_label}.")
        row = kv_swatch(row, "-- (red)", "\u2014",
                        FILLS["unavail"], FONTS["unavail"],
                        "Symbol unavailable on this C++ version.")
        row = kv_swatch(row, "-- (grey)", "\u2014",
                        FILLS["not_cx"], FONTS["not_cx"],
                        "In the constexpr block: symbol exists at this tier "
                        "but is not constexpr.")
    if cfg["columns"].get("intrinsic", False):
        row = kv_swatch(row, "Yes (intrinsic)", "Yes",
                        FILLS["intrinsic"], FONTS["intrinsic"],
                        "Symbol depends on a compiler builtin.")


# ============================================================================
# To Do sheet
# ============================================================================

def _completeness(cfg, syms):
    """Two metrics: floor-coverage (how many reach the language floor) and
    shipped-completeness (how many have any implementation). The first
    drives the text; the second drives the colour."""
    floor = cfg["versions"][0]
    has_secondary = cfg["tiers"].get("secondary") is not None
    total = len(syms)
    if has_secondary:
        at_floor = sum(1 for s in syms if s.secondary_min == floor)
        shipped  = sum(1 for s in syms if s.secondary_min)
    else:
        at_floor = sum(1 for s in syms if s.primary_in == floor)
        shipped  = sum(1 for s in syms if s.primary_in)
    floor_pct   = (at_floor / total * 100.0) if total else 0.0
    shipped_pct = (shipped  / total * 100.0) if total else 0.0
    return at_floor, total, floor_pct, shipped, shipped_pct


def _todo_fill(shipped_pct, shipped_any, has_data):
    if not has_data or not shipped_any: return FILLS["todo_none"]
    if shipped_pct >= 100.0: return FILLS["todo_done"]
    if shipped_pct >=  75.0: return FILLS["todo_high"]
    if shipped_pct >=  50.0: return FILLS["todo_mid"]
    if shipped_pct >=  25.0: return FILLS["todo_low"]
    if shipped_pct >    0.0: return FILLS["todo_started"]
    return FILLS["todo_none"]


def _dep_satisfied(dep_str, headers):
    name = dep_str.strip().lstrip("<").rstrip(">").strip()
    for prefix in ("restd::", "djinterp::"):
        if name.startswith(prefix):
            name = name[len(prefix):]
    if not name: return False
    for hdr, syms in headers.items():
        hdr_clean = hdr.strip("<>").replace(".hpp", "")
        if name == hdr_clean or name == hdr:
            return any(s.primary_in or s.secondary_min for s in syms)
        for s in syms:
            if s.name == dep_str or s.name == name:
                return (s.primary_in is not None) or (s.secondary_min is not None)
    return False


def _blocked_symbols(headers):
    out = []
    for hdr, syms in headers.items():
        for s in syms:
            if not s.depends_on: continue
            total = len(s.depends_on)
            sat   = sum(1 for d in s.depends_on if _dep_satisfied(d, headers))
            miss  = [d for d in s.depends_on if not _dep_satisfied(d, headers)]
            out.append((hdr, s, sat, total, miss))
    out.sort(key=lambda r: (-(r[2] / r[3]) if r[3] else 0, r[0], r[1].name))
    return out


def build_todo_sheet(wb, cfg, headers, roadmap):
    ws = wb.create_sheet("To Do")
    write_title_bar(
        ws, "To Do",
        f"Per-{cfg['unit_label'].lower()} completeness summary, sorted "
        "least-complete first. Colour indicates floor coverage percentage. "
        "Bottom: blocked symbols sorted by ease-to-ship-now.",
        5)

    widths = [10, 24, 22, 80, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 4
    headings = ["Priority", cfg["unit_label"], "Coverage",
                "Summary / blockers", "Est. min C++"]
    for i, label in enumerate(headings, 1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = FONTS["header"]; c.fill = FILLS["header_fill"]
        c.alignment = ALIGN_CENTER; c.border = BORDER_ALL_THIN
    ws.row_dimensions[row].height = ROW_HEIGHTS["leaf_hdr"]
    row += 1

    json_headers     = set(headers.keys())
    roadmap_by_hdr   = {}
    for it in roadmap:
        roadmap_by_hdr.setdefault(it["header"], []).append(it)

    shipped_entries = []
    for hdr in sorted(json_headers):
        syms = headers[hdr]
        at_floor, total, floor_pct, shipped, shipped_pct = _completeness(cfg, syms)
        rm_entries = roadmap_by_hdr.get(hdr, [])
        rm = rm_entries[0] if rm_entries else None
        shipped_entries.append({
            "header": hdr, "c98": at_floor, "total": total,
            "c98_pct": floor_pct, "shipped": shipped,
            "shipped_pct": shipped_pct, "shipped_any": shipped > 0,
            "has_data": True,
            "priority": rm["priority"] if rm else 0,
            "summary":  rm["summary"]  if rm else "SHIPPED.",
            "est_min":  rm["est_min_cpp"] if rm else "",
            "is_meta":  False,
        })

    planned_entries = []
    for it in roadmap:
        if it["header"] in json_headers: continue
        planned_entries.append({
            "header": it["header"], "c98": 0, "total": 0,
            "c98_pct": 0.0, "shipped": 0, "shipped_pct": 0.0,
            "shipped_any": False, "has_data": False,
            "priority": it["priority"], "summary": it["summary"],
            "est_min": it["est_min_cpp"],
            "is_meta": (it["header"] == "(meta)"),
        })

    shipped_entries.sort(key=lambda e: (e["shipped_pct"], e["header"]))
    planned_entries.sort(key=lambda e: (e["is_meta"], e["priority"]))
    entries = shipped_entries + planned_entries

    for e in entries:
        if e["total"]:
            cov_text = f"{e['c98']}/{e['total']} - {e['c98_pct']:.1f}%"
        else:
            cov_text = "planned"
        if e["is_meta"]:
            fill = FILLS["reminder_fill"]; font = FONTS["reminder"]
        else:
            fill = _todo_fill(e["shipped_pct"], e["shipped_any"],
                              e["has_data"])
            font = FONTS["body"]
        cells = [
            (str(e["priority"]),    ALIGN_CENTER),
            (e["header"],           ALIGN_LEFT_TOP),
            (cov_text,              ALIGN_CENTER),
            (e["summary"],          ALIGN_LEFT_TOP),
            (e["est_min"],          ALIGN_CENTER),
        ]
        for col_idx, (value, align) in enumerate(cells, 1):
            c = ws.cell(row=row, column=col_idx, value=value)
            c.font = font; c.alignment = align
            c.border = BORDER_ALL_THIN; c.fill = fill
        ws.row_dimensions[row].height = (ROW_HEIGHTS["reminder"]
                                          if e["is_meta"]
                                          else ROW_HEIGHTS["data"])
        row += 1

    ws.freeze_panes = "A5"
    if row > 5:
        ws.auto_filter.ref = f"A4:E{row - 1}"

    # Blocked-symbol triage section
    blocked = _blocked_symbols(headers)
    if blocked:
        row += 2
        ws.merge_cells(f"A{row}:E{row}")
        c = ws.cell(row=row, column=1,
                    value="Blocked symbols — sorted by 'ease to ship now' "
                          "(green = all deps satisfied, top priority)")
        c.font = FONTS["section"]; c.fill = FILLS["section_fill"]
        c.alignment = ALIGN_LEFT
        ws.row_dimensions[row].height = ROW_HEIGHTS["section"]
        row += 1

        for i, label in enumerate([cfg["unit_label"], "Symbol",
                                   "Deps satisfied", "Missing", "Notes"], 1):
            c = ws.cell(row=row, column=i, value=label)
            c.font = FONTS["header"]; c.fill = FILLS["header_fill"]
            c.alignment = ALIGN_CENTER; c.border = BORDER_ALL_THIN
        ws.row_dimensions[row].height = ROW_HEIGHTS["leaf_hdr"]
        row += 1

        for hdr, sym, sat, total, missing in blocked:
            ratio = (sat / total) if total else 0
            if   ratio >= 1.0: fill = FILLS["blocked_ready"]
            elif ratio >  0.0: fill = FILLS["blocked_partial"]
            else:              fill = FILLS["blocked_full"]
            cells = [
                (hdr,                                       ALIGN_LEFT_TOP),
                (sym.name,                                  ALIGN_LEFT_TOP),
                (f"{sat}/{total}",                          ALIGN_CENTER),
                (", ".join(missing) if missing else "(none)",
                                                            ALIGN_LEFT_TOP),
                (sym.notes or "",                           ALIGN_LEFT_TOP),
            ]
            for col_idx, (value, align) in enumerate(cells, 1):
                c = ws.cell(row=row, column=col_idx, value=value)
                c.font = FONTS["body"]; c.alignment = align
                c.border = BORDER_ALL_THIN; c.fill = fill
            ws.row_dimensions[row].height = ROW_HEIGHTS["data"]
            row += 1


# ============================================================================
# Pending Dependencies sheet
# ============================================================================

def build_pending_deps(wb, cfg, headers):
    ws = wb.create_sheet("Pending Dependencies")
    write_title_bar(
        ws, "Pending Dependencies",
        "Symbols whose full coverage is blocked on a module that has not "
        "yet shipped. Auto-derived from the depends_on field.",
        5)
    columns = [(cfg["unit_label"], 14), ("Symbol", 36), ("Category", 26),
               ("Missing dependency", 36), ("Notes", 60)]
    for i, (label, width) in enumerate(columns, 1):
        c = ws.cell(row=4, column=i, value=label)
        c.font = FONTS["header"]; c.fill = FILLS["header_fill"]
        c.alignment = ALIGN_CENTER; c.border = BORDER_ALL_THIN
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[4].height = ROW_HEIGHTS["leaf_hdr"]

    row_idx = 5
    for header_name, syms in headers.items():
        for sym in syms:
            if sym.depends_on:
                values = [header_name, sym.name, display_category(cfg, sym.group),
                          ", ".join(sym.depends_on), sym.notes or ""]
                for col, val in enumerate(values, 1):
                    c = ws.cell(row=row_idx, column=col, value=val)
                    c.font = FONTS["body"]; c.alignment = ALIGN_LEFT_TOP
                    c.border = BORDER_ALL_THIN
                    if row_idx % 2 == 1: c.fill = FILLS["zebra"]
                ws.row_dimensions[row_idx].height = ROW_HEIGHTS["data"]
                row_idx += 1
    ws.freeze_panes = "A5"
    if row_idx > 5:
        ws.auto_filter.ref = f"A4:E{row_idx - 1}"


# ============================================================================
# Coverage Failures sheet (global)
# ============================================================================

def build_coverage_failures(wb, cfg, headers):
    ws = wb.create_sheet("Coverage Failures")
    floor = cfg["versions"][0]
    primary_label = cfg["tiers"]["primary"]["label"]
    sec = cfg["tiers"].get("secondary")
    floor_label = f"{sec['label']} min" if sec else f"{primary_label} min"

    write_title_bar(
        ws, "Coverage Failures",
        f"Every symbol that does not reach full {floor} coverage. "
        f"Includes both partial coverage and unimplemented symbols.",
        6)
    columns = [(cfg["unit_label"], 14), ("Symbol", 36), ("Category", 26),
               (f"{primary_label} introduced", 16),
               (floor_label, 16),
               ("Failure reason", 80)]
    for i, (label, width) in enumerate(columns, 1):
        c = ws.cell(row=4, column=i, value=label)
        c.font = FONTS["header"]; c.fill = FILLS["header_fill"]
        c.alignment = ALIGN_CENTER; c.border = BORDER_ALL_THIN
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[4].height = ROW_HEIGHTS["leaf_hdr"]

    row_idx = 5
    for header_name, syms in headers.items():
        for sym in syms:
            if not _is_failure(cfg, sym): continue
            floor_value = (sym.secondary_min if sec else sym.primary_in) \
                          or "NOT IMPLEMENTED"
            values = [header_name, sym.name, display_category(cfg, sym.group),
                      sym.primary_in or "\u2014",
                      floor_value,
                      sym.failure_reason or ""]
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.font = FONTS["body"]
                c.alignment = ALIGN_CENTER if col in (4, 5) else ALIGN_LEFT_TOP
                c.border = BORDER_ALL_THIN
                if row_idx % 2 == 1: c.fill = FILLS["zebra"]
            ws.row_dimensions[row_idx].height = ROW_HEIGHTS["data"]
            row_idx += 1
    ws.freeze_panes = "A5"
    if row_idx > 5:
        ws.auto_filter.ref = f"A4:F{row_idx - 1}"


# ============================================================================
# Re-exports sheet (global)
# ============================================================================

def build_reexports(wb, cfg, headers):
    """Lists symbols flagged reexport=true: identity-preserving re-exports of
    runtime / ABI / compiler / language-provided std entities that cannot be
    portably reimplemented (restd contributes only the in-namespace alias).
    Distinct from back-ports (restd implements them) and from
    intrinsic_required traits (restd-owned, builtin-backed). Auto-derived from
    the reexport field; the sheet is omitted entirely when nothing is flagged.
    """
    prim = cfg["tiers"]["primary"]["label"]
    sec  = cfg["tiers"].get("secondary", {}).get("label", "")
    ws = wb.create_sheet("Re-exports")
    write_title_bar(
        ws, "Re-exports",
        f"Symbols {sec or 'the library'} surfaces as identity-preserving "
        f"re-exports of {prim} (using-declarations / typedef aliases): "
        f"runtime-, ABI-, compiler-, or language-provided entities that "
        f"cannot be portably reimplemented, so {sec or 'the library'}::X is "
        f"{prim}::X. Auto-derived from the reexport field.",
        5)
    columns = [(cfg["unit_label"], 14), ("Symbol", 40), ("Category", 24),
               ("Why it is re-exported (notes)", 78)]
    for i, (label, width) in enumerate(columns, 1):
        c = ws.cell(row=4, column=i, value=label)
        c.font = FONTS["header"]; c.fill = FILLS["header_fill"]
        c.alignment = ALIGN_CENTER; c.border = BORDER_ALL_THIN
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[4].height = ROW_HEIGHTS["leaf_hdr"]

    row_idx = 5
    for header_name, syms in headers.items():
        for sym in syms:
            if getattr(sym, "reexport", False):
                values = [header_name, sym.name,
                          display_category(cfg, sym.group), sym.notes or ""]
                for col, val in enumerate(values, 1):
                    c = ws.cell(row=row_idx, column=col, value=val)
                    c.font = FONTS["body"]; c.alignment = ALIGN_LEFT_TOP
                    c.border = BORDER_ALL_THIN
                    if row_idx % 2 == 1: c.fill = FILLS["zebra"]
                ws.row_dimensions[row_idx].height = ROW_HEIGHTS["data"]
                row_idx += 1
    ws.freeze_panes = "A5"
    if row_idx > 5:
        ws.auto_filter.ref = f"A4:D{row_idx - 1}"


# ============================================================================
# Orchestration
# ============================================================================

def build(data_dir, config=None):
    cfg = load_config(data_dir, overrides=config)
    headers, subtitles, legend_blurbs = load_data(data_dir)
    roadmap = load_roadmap(data_dir)
    layout = build_layout(cfg)

    wb = Workbook()
    wb.remove(wb.active)

    sorted_headers = sorted(headers.keys())
    header_blurbs  = [(h, legend_blurbs.get(h, "")) for h in sorted_headers]

    # Fixed-first sheets
    if cfg["sheets"].get("legend", True):
        build_legend_sheet(wb, cfg, header_blurbs)
    if cfg["sheets"].get("todo", True):
        build_todo_sheet(wb, cfg, headers, roadmap)
    # Per-header sheets, alphabetical
    for hdr in sorted_headers:
        build_data_sheet(wb, cfg, layout, hdr,
                         subtitles.get(hdr), headers[hdr])
    # Fixed-last sheets
    if cfg["sheets"].get("pending_deps", True):
        build_pending_deps(wb, cfg, headers)
    if cfg["sheets"].get("failures", True):
        build_coverage_failures(wb, cfg, headers)
    # Only emit the Re-exports sheet when something is actually flagged, so
    # projects that never mark re-exports don't get an empty sheet.
    if cfg["sheets"].get("reexports", True) and any(
            getattr(s, "reexport", False)
            for syms in headers.values() for s in syms):
        build_reexports(wb, cfg, headers)

    return wb


def main():
    args = parse_args()
    wb = build(args.data_dir)
    os.makedirs(os.path.dirname(os.path.abspath(args.output)) or ".",
                exist_ok=True)
    wb.save(args.output)
    print(f"Wrote {args.output}")
    print(f"  data:    {args.data_dir}")
    print(f"  sheets ({len(wb.sheetnames)}): {wb.sheetnames}")


if __name__ == "__main__":
    main()
