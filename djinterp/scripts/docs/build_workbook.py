"""
restd coverage workbook builder.

Reads one JSON file per std header from <root>/docs/restd/data/<header>.json
and writes the workbook to <root>/docs/restd/restd_coverage.xlsx.

Project layout (paths relative to this script's location, scripts/restd/):
    ../../docs/restd/data/         -- per-header JSON files (source of truth)
    ../../docs/restd/data/_roadmap.json
    ../../docs/restd/restd_coverage.xlsx   (output)

Run:  python3 scripts/restd/build_workbook.py
  (or from anywhere — paths are resolved relative to this file via __file__)

The JSON layer is the single source of truth. Everything visual lives
in the STYLE section near the top of this file — colors, column widths,
row heights, fonts, sheet ordering — change any of those constants and
the whole workbook re-themes.
"""

import glob
import json
import os
from collections import namedtuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# Project paths
# ============================================================================

_HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.normpath(os.path.join(_HERE, "..", ".."))
DATA_DIR     = os.path.join(PROJECT_ROOT, "docs", "restd", "data")
OUTPUT_PATH  = os.path.join(PROJECT_ROOT, "docs", "restd", "restd_coverage.xlsx")

# ============================================================================
# Symbol record — mirrors the JSON schema.
# ============================================================================

Symbol = namedtuple("Symbol", [
    "name", "group",
    "std_in", "restd_min",
    "constexpr_in_std", "constexpr_in_restd",
    "intrinsic_required", "intrinsic_names", "detection_macro",
    "t_alias_in", "v_var_in", "deprecated_in",
    "notes", "depends_on", "failure_reason",
])

def symbol_from_dict(d):
    return Symbol(
        name               = d["name"],
        group              = d["group"],
        std_in             = d.get("std_in"),
        restd_min          = d.get("restd_min"),
        constexpr_in_std   = d.get("constexpr_in_std"),
        constexpr_in_restd = d.get("constexpr_in_restd"),
        intrinsic_required = bool(d.get("intrinsic_required", False)),
        intrinsic_names    = d.get("intrinsic_names", "") or "",
        detection_macro    = d.get("detection_macro", "") or "",
        t_alias_in         = d.get("t_alias_in"),
        v_var_in           = d.get("v_var_in"),
        deprecated_in      = d.get("deprecated_in"),
        notes              = d.get("notes", "") or "",
        depends_on         = tuple(d.get("depends_on") or ()),
        failure_reason     = d.get("failure_reason"),
    )


# ============================================================================
# Data loader
# ============================================================================

def load_data(data_dir):
    headers       = {}
    subtitles     = {}
    legend_blurbs = {}
    for path in sorted(glob.glob(os.path.join(data_dir, "*.json"))):
        if os.path.basename(path).startswith("_"):
            continue
        with open(path) as f:
            obj = json.load(f)
        h = obj["header"]
        headers[h]       = [symbol_from_dict(s) for s in obj["symbols"]]
        subtitles[h]     = obj.get("subtitle", "")
        legend_blurbs[h] = obj.get("legend_blurb", "")
    return headers, subtitles, legend_blurbs


def load_roadmap(data_dir):
    """Loads data/_roadmap.json -> list of dicts (kept as dicts, not tuples,
    so future fields can be added without breaking the builder)."""
    path = os.path.join(data_dir, "_roadmap.json")
    with open(path) as f:
        return json.load(f)


# ============================================================================
# STYLE — every visual setting lives in this section.
# Change anything here and the workbook re-themes consistently.
# ============================================================================

# ---- color palette (hex strings, no '#') ----
COLORS = {
    # data-cell fills
    "std":              "C6EFCE",   # green: std support
    "restd":            "FFEB9C",   # yellow: restd back-port
    "unavail":          "FFC7CE",   # red: unavailable
    "not_cx":           "E7E6E6",   # grey: exists but not constexpr
    "intrinsic":        "FCE4D6",   # peach: intrinsic-backed

    # header / structural fills
    "header_fill":      "305496",   # dark blue: leaf header row 5
    "subheader_fill":   "8EA9DB",   # mid blue: group header row 4
    "section_fill":     "DDEBF7",   # light blue: per-sheet section banners
    "zebra":            "F8F9FA",   # near-white: alternating data row
    "reminder_fill":    "FFF2CC",   # cream: meta reminder row

    # To Do sheet completeness gradient (red -> yellow -> green)
    "todo_done":        "70AD47",   # 100% complete
    "todo_high":        "C6EFCE",   # 75-99%
    "todo_mid":         "FFEB9C",   # 50-74%
    "todo_low":         "FFD580",   # 25-49%
    "todo_started":     "F4A582",   # 1-24%
    "todo_none":        "F8CBAD",   # 0% (or planned/not-started)

    # Blocked-symbol triage gradient
    "blocked_ready":    "C6EFCE",   # all deps satisfied -> top priority
    "blocked_partial":  "FFEB9C",   # some deps satisfied
    "blocked_full":     "FFC7CE",   # no deps satisfied -> low priority

    # text colors
    "text_std":         "006100",
    "text_restd":       "9C5700",
    "text_unavail":     "9C0006",
    "text_not_cx":      "595959",
    "text_intrinsic":   "9C5700",
    "text_header":      "FFFFFF",
    "text_title":       "305496",
    "text_subtitle":    "595959",
    "text_body":        "000000",
    "text_reminder":    "7F6000",
    "text_section":     "305496",
}

# ---- typography ----
FONT_FAMILY = "Calibri"
FONT_SIZES = {
    "title":     18,
    "section":   11,
    "header":    11,
    "subheader": 10,
    "body":      10,
    "subtitle":  10,
}

# ---- column widths (per-header data sheet) ----
# Versioned columns (E-R) are uniform; everything else is named explicitly.
COLUMN_WIDTHS = {
    "std_header":      12,    # A
    "symbol":          34,    # B
    "category":        26,    # C
    "added_in":        12,    # D
    "version":          8,    # E-R: uniform 8-wide per spec
    "t_alias":         14,    # S
    "v_var":           14,    # T
    "deprecated":      14,    # U
    "intrinsic":       13,    # V
    "intrinsic_names": 34,    # W
    "detection_macro": 38,    # X
    "notes":           70,    # Y
}

# ---- row heights ----
ROW_HEIGHTS = {
    "title":       26,
    "subtitle":    18,
    "spacer":       6,
    "group_hdr":   22,   # row 4 ("coverage", "constexpr in")
    "leaf_hdr":    28,   # row 5 (C++98, C++11, ..., leaf labels)
    "data":        30,
    "section":     22,   # per-sheet section banners
    "reminder":    36,
}

# ---- border styles ----
# Thin borders between every cell; thicker borders at column-group boundaries.
BORDER_THIN_COLOR  = "BFBFBF"
BORDER_THICK_COLOR = "305496"
BORDER_THICK_STYLE = "medium"

# ---- sheet ordering ----
# Fixed-position sheets at the start and end; per-header sheets fill
# the middle in alphabetical order.
SHEET_ORDER_FIRST = ["Legend", "To Do"]
SHEET_ORDER_LAST  = ["Pending Dependencies", "Coverage Failures"]


# ============================================================================
# STRUCTURAL CONSTANTS (column layout, version axis)
# ============================================================================

VERSIONS = ["C++98", "C++11", "C++14", "C++17", "C++20", "C++23", "C++26"]

# Column slots (used by the data-sheet builder).
# slot_key, leaf_label, width_key, leaf_alignment
# Note: the version columns (cov_*, cx_*) use a special left-aligned leaf
# header (per spec) but center-aligned data cells.
COLUMN_LEAVES = [
    # slot              leaf label                    width_key         leaf align
    ("std_header",      "Std Header",                 "std_header",     "left"),
    ("symbol",          "Symbol",                     "symbol",         "left"),
    ("category",        "Category",                   "category",       "left"),
    ("added_in",        "std",                        "added_in",       "center"),
    ("cov_c98",         "C++98",                      "version",        "left"),
    ("cov_c11",         "C++11",                      "version",        "left"),
    ("cov_c14",         "C++14",                      "version",        "left"),
    ("cov_c17",         "C++17",                      "version",        "left"),
    ("cov_c20",         "C++20",                      "version",        "left"),
    ("cov_c23",         "C++23",                      "version",        "left"),
    ("cov_c26",         "C++26",                      "version",        "left"),
    ("cx_c98",          "C++98",                      "version",        "left"),
    ("cx_c11",          "C++11",                      "version",        "left"),
    ("cx_c14",          "C++14",                      "version",        "left"),
    ("cx_c17",          "C++17",                      "version",        "left"),
    ("cx_c20",          "C++20",                      "version",        "left"),
    ("cx_c23",          "C++23",                      "version",        "left"),
    ("cx_c26",          "C++26",                      "version",        "left"),
    ("t_alias",         "since",                      "t_alias",        "center"),
    ("v_var",           "since",                      "v_var",          "center"),
    ("deprecated",      "in",                         "deprecated",     "center"),
    ("intrinsic",       "required?",                  "intrinsic",      "center"),
    ("intrinsic_names", "Intrinsic(s)",               "intrinsic_names","left"),
    ("detection_macro", "restd detection macro",      "detection_macro","left"),
    ("notes",           "Notes / fallback behaviour", "notes",          "left"),
]

# Group-header (row 4) labels. Each tuple is (first_slot, last_slot, label).
# slots in the same group are merged across row 4. A label of None means
# the cells in row 4 are merged with their row-5 leaf header (no group
# label — the leaf label spans both rows).
COLUMN_GROUPS = [
    ("std_header",      "std_header",      None),
    ("symbol",          "symbol",          None),
    ("category",        "category",        None),
    ("added_in",        "added_in",        "Added to"),
    ("cov_c98",         "cov_c26",         "coverage"),
    ("cx_c98",          "cx_c26",          "constexpr in"),
    ("t_alias",         "t_alias",         "_t alias"),
    ("v_var",           "v_var",           "_v variable"),
    ("deprecated",      "deprecated",      "Deprecated"),
    ("intrinsic",       "intrinsic",       "Compiler intrinsic"),
    ("intrinsic_names", "intrinsic_names", None),
    ("detection_macro", "detection_macro", None),
    ("notes",           "notes",           None),
]

# Column slots whose RIGHT boundary gets a thicker border (group separator).
# Per spec: between D|E, K|L, R|S — i.e., after added_in, after cov_c26,
# after cx_c26.
GROUP_SEPARATORS_AFTER = {"added_in", "cov_c26", "cx_c26"}


# ============================================================================
# Derived openpyxl-typed style objects (built from the constants above)
# ============================================================================

def _fill(c):  return PatternFill("solid", fgColor=COLORS[c])
def _color(c): return COLORS[c]

FILLS = {k: _fill(k) for k in (
    "std", "restd", "unavail", "not_cx", "intrinsic",
    "header_fill", "subheader_fill", "section_fill", "zebra", "reminder_fill",
    "todo_done", "todo_high", "todo_mid", "todo_low", "todo_started", "todo_none",
    "blocked_ready", "blocked_partial", "blocked_full",
)}

def _font(name, size=None, bold=False, italic=False):
    return Font(
        name   = FONT_FAMILY,
        size   = size or FONT_SIZES["body"],
        color  = COLORS[name],
        bold   = bold,
        italic = italic,
    )

FONTS = {
    "std":       _font("text_std"),
    "restd":     _font("text_restd"),
    "unavail":   _font("text_unavail"),
    "not_cx":    _font("text_not_cx", italic=True),
    "intrinsic": _font("text_intrinsic", bold=True),
    "header":    _font("text_header",  size=FONT_SIZES["header"],    bold=True),
    "subheader": _font("text_header",  size=FONT_SIZES["subheader"], bold=True),
    "title":     _font("text_title",   size=FONT_SIZES["title"],     bold=True),
    "subtitle":  _font("text_subtitle",size=FONT_SIZES["subtitle"],  italic=True),
    "body":      _font("text_body"),
    "reminder":  _font("text_reminder",bold=True),
    "section":   _font("text_section", size=FONT_SIZES["section"],   bold=True),
}

SIDE_THIN  = Side(style="thin",              color=BORDER_THIN_COLOR)
SIDE_THICK = Side(style=BORDER_THICK_STYLE,  color=BORDER_THICK_COLOR)

BORDER_ALL_THIN = Border(top=SIDE_THIN, bottom=SIDE_THIN, left=SIDE_THIN, right=SIDE_THIN)

ALIGN_CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_LEFT_TOP = Alignment(horizontal="left",   vertical="top",    wrap_text=True)


# ============================================================================
# Cell-classification logic (data cells: support and constexpr)
# ============================================================================

def _ge(v_a, v_b):
    if v_a is None or v_b is None:
        return False
    return VERSIONS.index(v_a) >= VERSIONS.index(v_b)

def support_cell(version, std_in, restd_min):
    if std_in and _ge(version, std_in):
        return ("std",     FILLS["std"],     FONTS["std"])
    if restd_min and _ge(version, restd_min):
        return ("restd",   FILLS["restd"],   FONTS["restd"])
    return ("\u2014",      FILLS["unavail"], FONTS["unavail"])

def constexpr_cell(version, std_in, restd_min, cx_std, cx_restd):
    exists = (std_in and _ge(version, std_in)) or (restd_min and _ge(version, restd_min))
    if not exists:
        return ("\u2014",  FILLS["unavail"], FONTS["unavail"])
    if cx_std and _ge(version, cx_std):
        return ("cx",      FILLS["std"],     FONTS["std"])
    if cx_restd and _ge(version, cx_restd):
        return ("cx",      FILLS["restd"],   FONTS["restd"])
    return ("\u2014",      FILLS["not_cx"],  FONTS["not_cx"])


# ============================================================================
# Per-cell border helper — thin borders, thick at group separators.
# ============================================================================

# Map slot_key -> column index (1-based) for the data sheet layout.
SLOT_TO_COL = {leaf[0]: i + 1 for i, leaf in enumerate(COLUMN_LEAVES)}
COL_TO_SLOT = {i + 1: leaf[0] for i, leaf in enumerate(COLUMN_LEAVES)}

# Column indices that have a thick RIGHT border (group separator).
THICK_RIGHT_COLS = {SLOT_TO_COL[k] for k in GROUP_SEPARATORS_AFTER}

def data_cell_border(col_idx):
    right = SIDE_THICK if col_idx in THICK_RIGHT_COLS else SIDE_THIN
    return Border(top=SIDE_THIN, bottom=SIDE_THIN, left=SIDE_THIN, right=right)


# ============================================================================
# Generic write helpers
# ============================================================================

def write_title_bar(ws, title, subtitle, col_count):
    end_col = get_column_letter(col_count)
    ws.merge_cells(f"A1:{end_col}1")
    ws["A1"] = title
    ws["A1"].font      = FONTS["title"]
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = ROW_HEIGHTS["title"]
    ws.merge_cells(f"A2:{end_col}2")
    ws["A2"] = subtitle
    ws["A2"].font      = FONTS["subtitle"]
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = ROW_HEIGHTS["subtitle"]
    ws.row_dimensions[3].height = ROW_HEIGHTS["spacer"]


def set_column_widths():
    """Returns a callable that applies the standard column widths to a sheet."""
    def _apply(ws):
        for slot, leaf, width_key, _align in COLUMN_LEAVES:
            col_letter = get_column_letter(SLOT_TO_COL[slot])
            ws.column_dimensions[col_letter].width = COLUMN_WIDTHS[width_key]
    return _apply


# ============================================================================
# Header sheet — the per-std-header data sheet
# ============================================================================
#
# Layout (rows):
#   1: title bar
#   2: subtitle
#   3: spacer
#   4: group header (merged "coverage", "constexpr in", etc)
#   5: leaf header (Std Header, Symbol, C++98, C++11, ...)
#   6+: data
#   N+spacer: per-sheet Coverage Failures section (if any)
# ============================================================================

def _write_header_block(ws):
    """Writes rows 4 and 5 — group headers and leaf headers."""
    # Width setup
    set_column_widths()(ws)

    # Row 4 — group headers
    for first_slot, last_slot, label in COLUMN_GROUPS:
        first = SLOT_TO_COL[first_slot]
        last  = SLOT_TO_COL[last_slot]
        first_letter = get_column_letter(first)
        last_letter  = get_column_letter(last)

        if label is None:
            # Merge row 4 with row 5 for a single-spanning leaf header.
            ws.merge_cells(f"{first_letter}4:{last_letter}5")
            leaf_label = next(l[1] for l in COLUMN_LEAVES if l[0] == first_slot)
            c = ws.cell(row=4, column=first, value=leaf_label)
            c.font      = FONTS["header"]
            c.fill      = FILLS["header_fill"]
            c.alignment = ALIGN_CENTER
        else:
            # Merge across row 4 only for the group label.
            if first != last:
                ws.merge_cells(start_row=4, end_row=4, start_column=first, end_column=last)
            c = ws.cell(row=4, column=first, value=label)
            c.font      = FONTS["subheader"]
            c.fill      = FILLS["subheader_fill"]
            c.alignment = ALIGN_CENTER

    # Row 5 — leaf headers (one per column for groups whose row 4 has a label).
    # Skip slots that already had row 4 merged with row 5.
    groups_with_labels = {first for first, _, label in COLUMN_GROUPS if label is not None
                          for first in (first,)}  # placeholder; use real iteration below
    groups_with_labels = set()
    for first_slot, last_slot, label in COLUMN_GROUPS:
        if label is not None:
            # Every slot from first..last (by index) gets a row-5 leaf header.
            first = SLOT_TO_COL[first_slot]
            last  = SLOT_TO_COL[last_slot]
            for col in range(first, last + 1):
                slot_at_col = COL_TO_SLOT[col]
                leaf_label  = next(l[1] for l in COLUMN_LEAVES if l[0] == slot_at_col)
                c = ws.cell(row=5, column=col, value=leaf_label)
                c.font      = FONTS["header"]
                c.fill      = FILLS["header_fill"]
                # version columns: left-aligned in row 5 only; others: centered
                _slot_align = next(l[3] for l in COLUMN_LEAVES if l[0] == slot_at_col)
                c.alignment = ALIGN_LEFT if _slot_align == "left" else ALIGN_CENTER

    # Apply borders to both header rows (thin everywhere, thick at separators).
    for row in (4, 5):
        for col in range(1, len(COLUMN_LEAVES) + 1):
            ws.cell(row=row, column=col).border = data_cell_border(col)

    ws.row_dimensions[4].height = ROW_HEIGHTS["group_hdr"]
    ws.row_dimensions[5].height = ROW_HEIGHTS["leaf_hdr"]


def build_data_sheet(wb, header_name, subtitle_override, symbols):
    ws = wb.create_sheet(header_name)
    n_cols = len(COLUMN_LEAVES)

    write_title_bar(
        ws,
        header_name,
        f"Module surface: {len(symbols)} symbols. "
        "Green = std support. Yellow = restd back-port. Red = unavailable. "
        "The constexpr block tracks the same axis but for compile-time evaluation.",
        n_cols,
    )

    _write_header_block(ws)

    # Data rows start at row 6
    DATA_START = 6
    for row_idx, sym in enumerate(symbols, start=DATA_START):
        zebra = (row_idx % 2 == 1)

        def _set(col, value, font=FONTS["body"], align=ALIGN_LEFT_TOP, fill_zebra=True):
            c = ws.cell(row=row_idx, column=col, value=value)
            c.font      = font
            c.alignment = align
            c.border    = data_cell_border(col)
            if fill_zebra and zebra:
                c.fill = FILLS["zebra"]
            return c

        _set(SLOT_TO_COL["std_header"], header_name if row_idx == DATA_START else None)
        _set(SLOT_TO_COL["symbol"],     sym.name)
        _set(SLOT_TO_COL["category"],   sym.group)
        _set(SLOT_TO_COL["added_in"],   sym.std_in or "\u2014", align=ALIGN_CENTER)

        for j, ver in enumerate(VERSIONS):
            col = SLOT_TO_COL["cov_c98"] + j
            text, fill, font = support_cell(ver, sym.std_in, sym.restd_min)
            c = ws.cell(row=row_idx, column=col, value=text)
            c.font = font; c.fill = fill; c.alignment = ALIGN_CENTER
            c.border = data_cell_border(col)

        for j, ver in enumerate(VERSIONS):
            col = SLOT_TO_COL["cx_c98"] + j
            text, fill, font = constexpr_cell(
                ver, sym.std_in, sym.restd_min,
                sym.constexpr_in_std, sym.constexpr_in_restd,
            )
            c = ws.cell(row=row_idx, column=col, value=text)
            c.font = font; c.fill = fill; c.alignment = ALIGN_CENTER
            c.border = data_cell_border(col)

        _set(SLOT_TO_COL["t_alias"],    sym.t_alias_in    or "\u2014", align=ALIGN_CENTER)
        _set(SLOT_TO_COL["v_var"],      sym.v_var_in      or "\u2014", align=ALIGN_CENTER)
        _set(SLOT_TO_COL["deprecated"], sym.deprecated_in or "\u2014", align=ALIGN_CENTER)

        if sym.intrinsic_required:
            c = ws.cell(row=row_idx, column=SLOT_TO_COL["intrinsic"], value="Yes")
            c.font = FONTS["intrinsic"]; c.fill = FILLS["intrinsic"]
            c.alignment = ALIGN_CENTER; c.border = data_cell_border(SLOT_TO_COL["intrinsic"])
        else:
            _set(SLOT_TO_COL["intrinsic"], "No", align=ALIGN_CENTER)

        _set(SLOT_TO_COL["intrinsic_names"], sym.intrinsic_names or "\u2014")
        _set(SLOT_TO_COL["detection_macro"], sym.detection_macro or "\u2014")
        _set(SLOT_TO_COL["notes"],            sym.notes or "")

        ws.row_dimensions[row_idx].height = ROW_HEIGHTS["data"]

    ws.freeze_panes   = "E6"
    ws.auto_filter.ref = f"A5:{get_column_letter(n_cols)}{DATA_START - 1 + len(symbols)}"

    # ----------------------------------------------------------------------
    # Per-sheet local Coverage Failures section.
    # Shows symbols from THIS header that don't reach full C++98 coverage,
    # with their failure_reason.
    # ----------------------------------------------------------------------
    failures = [s for s in symbols if _is_failure(s)]
    if failures:
        _write_local_failures(ws, header_name, failures, DATA_START + len(symbols))


def _is_failure(sym):
    return (
        sym.failure_reason is not None
        or sym.restd_min is None
        or (sym.restd_min and sym.restd_min != "C++98")
    )


def _write_local_failures(ws, header_name, failures, start_row):
    """Mini failures table appended below the main data table."""
    row = start_row + 2  # spacer

    # Section banner — merge across A..E (left half of the sheet).
    n_cols = len(COLUMN_LEAVES)
    banner_end = get_column_letter(n_cols)
    ws.merge_cells(f"A{row}:{banner_end}{row}")
    c = ws.cell(row=row, column=1,
                value=f"Coverage Failures (this sheet only) — "
                      f"{len(failures)} symbol(s) below 100% C++98 coverage")
    c.font      = FONTS["section"]
    c.fill      = FILLS["section_fill"]
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[row].height = ROW_HEIGHTS["section"]

    row += 1

    # Mini header row
    mini_cols = [
        ("Symbol",        SLOT_TO_COL["symbol"]),
        ("Category",      SLOT_TO_COL["category"]),
        ("Std introduced",SLOT_TO_COL["added_in"]),
        ("restd min",     SLOT_TO_COL["cov_c98"]),
        ("Failure reason",SLOT_TO_COL["intrinsic_names"]),  # spans the wide notes columns
    ]
    # Header cells
    layout = [
        (1, SLOT_TO_COL["symbol"],          "Symbol"),
        (SLOT_TO_COL["symbol"] + 1, SLOT_TO_COL["category"], None),  # leave gap merge
    ]
    # Simpler: write 5 columns A..E reusing the main column slots for width.
    headings = ["Symbol", "Category", "Std introduced", "restd min", "Failure reason"]
    starts   = [
        1,
        SLOT_TO_COL["symbol"] + 1,
        SLOT_TO_COL["category"] + 1,
        SLOT_TO_COL["added_in"] + 1,
        SLOT_TO_COL["cov_c98"] + 1,
    ]
    ends     = [
        SLOT_TO_COL["symbol"],
        SLOT_TO_COL["category"],
        SLOT_TO_COL["added_in"],
        SLOT_TO_COL["cov_c98"],
        n_cols,
    ]
    for label, s, e in zip(headings, starts, ends):
        if s != e:
            ws.merge_cells(start_row=row, end_row=row, start_column=s, end_column=e)
        c = ws.cell(row=row, column=s, value=label)
        c.font      = FONTS["header"]
        c.fill      = FILLS["header_fill"]
        c.alignment = ALIGN_CENTER
        c.border    = BORDER_ALL_THIN
    ws.row_dimensions[row].height = ROW_HEIGHTS["leaf_hdr"]
    row += 1

    # Data rows
    for i, sym in enumerate(failures):
        zebra = (i % 2 == 0)
        cells_data = [
            (sym.name,                         starts[0], ends[0], ALIGN_LEFT_TOP),
            (sym.group,                        starts[1], ends[1], ALIGN_LEFT_TOP),
            (sym.std_in or "\u2014",           starts[2], ends[2], ALIGN_CENTER),
            (sym.restd_min or "NOT IMPLEMENTED", starts[3], ends[3], ALIGN_CENTER),
            (sym.failure_reason or "",         starts[4], ends[4], ALIGN_LEFT_TOP),
        ]
        for value, s, e, align in cells_data:
            if s != e:
                ws.merge_cells(start_row=row, end_row=row, start_column=s, end_column=e)
            c = ws.cell(row=row, column=s, value=value)
            c.font      = FONTS["body"]
            c.alignment = align
            c.border    = BORDER_ALL_THIN
            if zebra:
                c.fill = FILLS["zebra"]
        ws.row_dimensions[row].height = ROW_HEIGHTS["data"]
        row += 1


# ============================================================================
# Legend sheet
# ============================================================================

LEGEND_HEADER_ROW = ("Legend", "You are here. Colour key + column meanings.")

AUX_SHEET_BLURBS = [
    ("To Do",                 "Per-header completeness summary + dependency triage. Replaces the old Roadmap sheet."),
    ("Pending Dependencies",  "Symbols that need other restd modules before they reach full coverage. Auto-derived from the depends_on field."),
    ("Coverage Failures",     "Detailed list of every symbol that does not reach C++98+ coverage, with the reason."),
]

LEGEND_COLUMN_MEANINGS = [
    ("Added to std",                 "C++ version when the symbol entered std::."),
    ("coverage cells (7)",           "Support tier across C++98 -> C++26."),
    ("constexpr cells (7)",          "Constexpr tier on the same axis: green = constexpr in std, yellow = constexpr in restd ahead of std, grey = exists but not constexpr, red = unavailable."),
    ("_t alias since",               "C++ version when std::<symbol>_t alias was added."),
    ("_v variable since",            "C++ version when std::<symbol>_v variable was added."),
    ("Deprecated in",                "C++ version when the symbol was deprecated; -- when active."),
    ("Compiler intrinsic required?", "Yes when the symbol depends on a compiler builtin like __is_class, __is_enum, __underlying_type, alignof. Such rows show the intrinsic and a D_RESTD_HAS_* override macro."),
    ("Intrinsic(s)",                 "Compiler builtin(s) used by both std and restd; -- when none."),
    ("restd detection macro",        "Predefinable macro that resolves to 1/0 based on intrinsic availability. Override before #include to force."),
    ("Notes / fallback behaviour",   "Behaviour when the trait is gated, intrinsic is missing, or the implementation has known limitations."),
]

LEGEND_LABEL_DESCRIPTIONS = {
    "std":                "Symbol is part of std at this C++ version. Use std directly.",
    "restd":              "restd back-ports the symbol on this C++ version (std does not yet have it).",
    "cx (green)":         "In the constexpr block: symbol is constexpr in std at this version.",
    "cx (yellow)":        "In the constexpr block: symbol is constexpr in restd ahead of std.",
    "-- (red, support)":  "Symbol unavailable on this C++ version.",
    "-- (grey, constexpr)":"In the constexpr block: symbol exists at this tier but is not constexpr.",
    "Yes (intrinsic)":    "In 'Compiler intrinsic required?': this symbol depends on a compiler builtin.",
}


def build_legend_sheet(wb, header_blurbs):
    ws = wb.create_sheet("Legend")
    write_title_bar(
        ws,
        "restd module coverage matrix",
        "Single source of truth: data/*.json. Builder: build_workbook.py. "
        "All visual settings live in the STYLE section of build_workbook.py.",
        4,
    )

    def section(row, title):
        c = ws.cell(row=row, column=1, value=title)
        c.font = FONTS["section"]
        ws.row_dimensions[row].height = ROW_HEIGHTS["section"]
        return row + 1

    def kv(row, label, descr):
        a = ws.cell(row=row, column=1, value=label)
        a.font = FONTS["body"]; a.alignment = ALIGN_LEFT_TOP; a.border = BORDER_ALL_THIN
        b = ws.cell(row=row, column=2, value=descr)
        b.font = FONTS["body"]; b.alignment = ALIGN_LEFT_TOP; b.border = BORDER_ALL_THIN
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 28
        return row + 1

    def kv_swatch(row, label, sample_text, fill, font):
        a = ws.cell(row=row, column=1, value=label)
        a.font = FONTS["body"]; a.alignment = ALIGN_LEFT_TOP; a.border = BORDER_ALL_THIN
        sw = ws.cell(row=row, column=2, value=sample_text)
        sw.fill = fill; sw.font = font; sw.alignment = ALIGN_CENTER; sw.border = BORDER_ALL_THIN
        b = ws.cell(row=row, column=3, value=LEGEND_LABEL_DESCRIPTIONS[label])
        b.font = FONTS["body"]; b.alignment = ALIGN_LEFT_TOP; b.border = BORDER_ALL_THIN
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 28
        return row + 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 30

    row = 4
    row = section(row, "Sheets")
    row = kv(row, *LEGEND_HEADER_ROW)
    for sheet_name, descr in header_blurbs:
        row = kv(row, sheet_name, descr)
    for sheet_name, descr in AUX_SHEET_BLURBS:
        row = kv(row, sheet_name, descr)

    row += 1
    row = section(row, "Colour key")
    row = kv_swatch(row, "std",                  "std",   FILLS["std"],       FONTS["std"])
    row = kv_swatch(row, "restd",                "restd", FILLS["restd"],     FONTS["restd"])
    row = kv_swatch(row, "cx (green)",           "cx",    FILLS["std"],       FONTS["std"])
    row = kv_swatch(row, "cx (yellow)",          "cx",    FILLS["restd"],     FONTS["restd"])
    row = kv_swatch(row, "-- (red, support)",    "\u2014", FILLS["unavail"], FONTS["unavail"])
    row = kv_swatch(row, "-- (grey, constexpr)", "\u2014", FILLS["not_cx"],  FONTS["not_cx"])
    row = kv_swatch(row, "Yes (intrinsic)",      "Yes",   FILLS["intrinsic"], FONTS["intrinsic"])

    row += 1
    row = section(row, "Other column meanings")
    for k, v in LEGEND_COLUMN_MEANINGS:
        row = kv(row, k, v)


# ============================================================================
# To Do sheet (replaces the old Roadmap sheet)
# ============================================================================

def _completeness(syms):
    """Returns (c98_count, total, c98_pct, shipped_count, shipped_pct).

    Two metrics:
      - C++98 coverage: how many symbols reach C++98? Used in the TEXT.
      - Shipped completeness: how many symbols have a restd_min set
        (i.e., are implemented at any tier)? Used for the COLOR gradient.

    This separation matters for modules with a hard language floor
    (e.g. <expected> is C++11+; its C++98 coverage is 0% but it can be
    100% shipped). Coloring by shipped-% gives a sensible 'is this
    done?' indicator while the text still shows the back-port reach.
    """
    total = len(syms)
    c98     = sum(1 for s in syms if s.restd_min == "C++98")
    shipped = sum(1 for s in syms if s.restd_min)
    c98_pct     = (c98 / total * 100.0)     if total else 0.0
    shipped_pct = (shipped / total * 100.0) if total else 0.0
    return c98, total, c98_pct, shipped, shipped_pct

def _todo_fill(shipped_pct, shipped_any, has_data):
    """Pick a completeness-graduated fill colour.

    shipped_pct drives the gradient: 100% = darkest green, 0% = peach/red.
    has_data=False means the module has no JSON file yet (planned only).
    """
    if not has_data or not shipped_any:
        return FILLS["todo_none"]
    if shipped_pct >= 100.0:
        return FILLS["todo_done"]
    if shipped_pct >= 75.0:
        return FILLS["todo_high"]
    if shipped_pct >= 50.0:
        return FILLS["todo_mid"]
    if shipped_pct >= 25.0:
        return FILLS["todo_low"]
    if shipped_pct > 0.0:
        return FILLS["todo_started"]
    return FILLS["todo_none"]


def _dep_satisfied(dep_str, headers):
    """Best-effort: is this dependency now shipped somewhere?
    A dep is treated as satisfied if a header by that name has ANY shipped
    symbols, or if a symbol by that name appears in any shipped header.
    """
    # Normalize: strip <>, strip restd:: prefix.
    name = dep_str.strip()
    if name.startswith("<") and name.endswith(">"):
        name = name[1:-1]
    if name.startswith("restd::"):
        name = name[len("restd::"):]
    name = name.strip()
    if not name:
        return False
    # Header match
    for hdr, syms in headers.items():
        hdr_clean = hdr.strip("<>")
        if name == hdr_clean:
            return any(s.restd_min for s in syms)
        for s in syms:
            if s.name == dep_str or s.name == name:
                return s.restd_min is not None
    return False


def _blocked_symbols(headers):
    """Returns list of (header, symbol, satisfied_count, total_deps, missing_list)."""
    out = []
    for hdr, syms in headers.items():
        for s in syms:
            if not s.depends_on:
                continue
            total = len(s.depends_on)
            sat   = sum(1 for d in s.depends_on if _dep_satisfied(d, headers))
            missing = [d for d in s.depends_on if not _dep_satisfied(d, headers)]
            out.append((hdr, s, sat, total, missing))
    # Sort: highest ratio satisfied first (easiest to ship now).
    out.sort(key=lambda r: (-(r[2] / r[3]) if r[3] else 0, r[0], r[1].name))
    return out


def build_todo_sheet(wb, headers, roadmap):
    ws = wb.create_sheet("To Do")
    write_title_bar(
        ws,
        "To Do",
        "Per-header completeness summary, sorted least-complete first. "
        "Colour indicates C++98 coverage percentage. Bottom section: individual "
        "symbols whose full coverage is blocked on external modules — priority "
        "ordered by how many of their dependencies are now shipped.",
        5,
    )

    # Column widths
    widths = [10, 24, 22, 80, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Header row
    row = 4
    headings = ["Priority", "Header", "Coverage", "Summary / blockers", "Est. min C++"]
    for i, label in enumerate(headings, 1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = FONTS["header"]; c.fill = FILLS["header_fill"]
        c.alignment = ALIGN_CENTER; c.border = BORDER_ALL_THIN
    ws.row_dimensions[row].height = ROW_HEIGHTS["leaf_hdr"]
    row += 1

    # ----- Build the unified list of {header -> roadmap entry, completeness} -----
    # Headers that have JSON files
    json_headers = set(headers.keys())
    # Headers that appear in the roadmap (any priority)
    roadmap_headers = {it["header"] for it in roadmap}
    # Roadmap lookup by header
    roadmap_by_header = {}
    for it in roadmap:
        roadmap_by_header.setdefault(it["header"], []).append(it)

    # Compose entries for shipped headers (those with JSON files)
    shipped_entries = []
    for hdr in sorted(json_headers):
        syms = headers[hdr]
        c98, total, c98_pct, shipped, shipped_pct = _completeness(syms)
        shipped_any = (shipped > 0)
        # Use the highest-priority roadmap entry for this header (typically 0
        # if shipped, or the planned-work priority if some symbols pending).
        rm_entries = roadmap_by_header.get(hdr, [])
        rm = rm_entries[0] if rm_entries else None
        priority = rm["priority"] if rm else 0
        summary  = rm["summary"]  if rm else "SHIPPED."
        est_min  = rm["est_min_cpp"] if rm else ""
        shipped_entries.append({
            "header": hdr,
            "c98": c98, "total": total, "c98_pct": c98_pct,
            "shipped": shipped, "shipped_pct": shipped_pct,
            "shipped_any": shipped_any, "has_data": True,
            "priority": priority,
            "summary": summary, "est_min": est_min, "is_meta": False,
        })

    # Planned-only entries (in roadmap, no JSON file, non-meta)
    planned_entries = []
    for it in roadmap:
        if it["header"] in json_headers:
            continue
        is_meta = (it["header"] == "(meta)")
        planned_entries.append({
            "header": it["header"],
            "c98": 0, "total": 0, "c98_pct": 0.0,
            "shipped": 0, "shipped_pct": 0.0,
            "shipped_any": False, "has_data": False,
            "priority": it["priority"],
            "summary": it["summary"], "est_min": it["est_min_cpp"],
            "is_meta": is_meta,
        })

    # Sort: shipped entries first (least-shipped first, then alphabetical),
    # then planned (by priority), with the (meta) reminder pinned to the end.
    shipped_entries.sort(key=lambda e: (e["shipped_pct"], e["header"]))
    planned_entries.sort(key=lambda e: (e["is_meta"], e["priority"]))

    entries = shipped_entries + planned_entries

    # ----- Emit per-header rows -----
    for i, e in enumerate(entries):
        zebra = (i % 2 == 0)
        if e["total"]:
            cov_text = f"{e['c98']}/{e['total']} - {e['c98_pct']:.1f}%"
        else:
            cov_text = "planned"
        if e["is_meta"]:
            fill = FILLS["reminder_fill"]
            font = FONTS["reminder"]
        else:
            fill = _todo_fill(e["shipped_pct"], e["shipped_any"], e["has_data"])
            font = FONTS["body"]

        cells = [
            (str(e["priority"]),   ALIGN_CENTER),
            (e["header"],          ALIGN_LEFT_TOP),
            (cov_text,             ALIGN_CENTER),
            (e["summary"],         ALIGN_LEFT_TOP),
            (e["est_min"],         ALIGN_CENTER),
        ]
        for col_idx, (value, align) in enumerate(cells, 1):
            c = ws.cell(row=row, column=col_idx, value=value)
            c.font = font; c.alignment = align; c.border = BORDER_ALL_THIN
            c.fill = fill
        ws.row_dimensions[row].height = ROW_HEIGHTS["reminder"] if e["is_meta"] \
                                        else ROW_HEIGHTS["data"]
        row += 1

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:E{row - 1}"

    # ----------------------------------------------------------------------
    # Bottom section: blocked symbols triage.
    # ----------------------------------------------------------------------
    blocked = _blocked_symbols(headers)
    if blocked:
        row += 2
        ws.merge_cells(f"A{row}:E{row}")
        c = ws.cell(row=row, column=1,
                    value="Blocked symbols — sorted by 'ease to ship now' "
                          "(green = all deps satisfied, top priority)")
        c.font = FONTS["section"]; c.fill = FILLS["section_fill"]
        c.alignment = ALIGN_LEFT
        ws.row_dimensions[row].height = ROW_HEIGHTS["section"]
        row += 1

        # Mini header
        blocked_headings = ["Header", "Symbol", "Deps satisfied", "Missing", "Notes"]
        for i, label in enumerate(blocked_headings, 1):
            c = ws.cell(row=row, column=i, value=label)
            c.font = FONTS["header"]; c.fill = FILLS["header_fill"]
            c.alignment = ALIGN_CENTER; c.border = BORDER_ALL_THIN
        ws.row_dimensions[row].height = ROW_HEIGHTS["leaf_hdr"]
        row += 1

        for i, (hdr, sym, sat, total, missing) in enumerate(blocked):
            ratio = (sat / total) if total else 0
            if   ratio >= 1.0: fill = FILLS["blocked_ready"]
            elif ratio >  0.0: fill = FILLS["blocked_partial"]
            else:              fill = FILLS["blocked_full"]

            cells = [
                (hdr,                                   ALIGN_LEFT_TOP),
                (sym.name,                              ALIGN_LEFT_TOP),
                (f"{sat}/{total}",                      ALIGN_CENTER),
                (", ".join(missing) if missing else "(none)", ALIGN_LEFT_TOP),
                (sym.notes or "",                       ALIGN_LEFT_TOP),
            ]
            for col_idx, (value, align) in enumerate(cells, 1):
                c = ws.cell(row=row, column=col_idx, value=value)
                c.font = FONTS["body"]; c.alignment = align
                c.border = BORDER_ALL_THIN; c.fill = fill
            ws.row_dimensions[row].height = ROW_HEIGHTS["data"]
            row += 1


# ============================================================================
# Pending Dependencies sheet
# ============================================================================

def build_pending_deps(wb, headers):
    ws = wb.create_sheet("Pending Dependencies")
    write_title_bar(
        ws,
        "Pending Dependencies",
        "Symbols whose full coverage is blocked on a restd module that has not "
        "yet shipped. Auto-derived from the depends_on field in the per-header "
        "JSON files.",
        5,
    )
    columns = [("Std Header", 14), ("Symbol", 36), ("Category", 26),
               ("Missing dependency", 36), ("Notes", 60)]
    for i, (label, width) in enumerate(columns, 1):
        c = ws.cell(row=4, column=i, value=label)
        c.font = FONTS["header"]; c.fill = FILLS["header_fill"]
        c.alignment = ALIGN_CENTER; c.border = BORDER_ALL_THIN
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[4].height = ROW_HEIGHTS["leaf_hdr"]

    row_idx = 5
    for header_name, syms in headers.items():
        for sym in syms:
            if sym.depends_on:
                missing = ", ".join(sym.depends_on)
                values = [header_name, sym.name, sym.group, missing, sym.notes or ""]
                for col, val in enumerate(values, 1):
                    c = ws.cell(row=row_idx, column=col, value=val)
                    c.font = FONTS["body"]; c.alignment = ALIGN_LEFT_TOP
                    c.border = BORDER_ALL_THIN
                    if row_idx % 2 == 1: c.fill = FILLS["zebra"]
                ws.row_dimensions[row_idx].height = ROW_HEIGHTS["data"]
                row_idx += 1
    ws.freeze_panes = "A5"
    if row_idx > 5:
        ws.auto_filter.ref = f"A4:E{row_idx - 1}"


# ============================================================================
# Global Coverage Failures sheet
# ============================================================================

def build_coverage_failures(wb, headers):
    ws = wb.create_sheet("Coverage Failures")
    write_title_bar(
        ws,
        "Coverage Failures",
        "Every symbol that does not reach full C++98+ coverage. "
        "Includes both partial coverage (restd_min > C++98) and unimplemented symbols.",
        6,
    )
    columns = [("Std Header", 14), ("Symbol", 36), ("Category", 26),
               ("Std introduced", 14), ("restd min", 16), ("Failure reason", 80)]
    for i, (label, width) in enumerate(columns, 1):
        c = ws.cell(row=4, column=i, value=label)
        c.font = FONTS["header"]; c.fill = FILLS["header_fill"]
        c.alignment = ALIGN_CENTER; c.border = BORDER_ALL_THIN
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[4].height = ROW_HEIGHTS["leaf_hdr"]

    row_idx = 5
    for header_name, syms in headers.items():
        for sym in syms:
            if not _is_failure(sym):
                continue
            values = [header_name, sym.name, sym.group,
                      sym.std_in or "\u2014",
                      sym.restd_min or "NOT IMPLEMENTED",
                      sym.failure_reason or ""]
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.font = FONTS["body"]
                c.alignment = ALIGN_CENTER if col in (4, 5) else ALIGN_LEFT_TOP
                c.border = BORDER_ALL_THIN
                if row_idx % 2 == 1: c.fill = FILLS["zebra"]
            ws.row_dimensions[row_idx].height = ROW_HEIGHTS["data"]
            row_idx += 1
    ws.freeze_panes = "A5"
    if row_idx > 5:
        ws.auto_filter.ref = f"A4:F{row_idx - 1}"


# ============================================================================
# Orchestration
# ============================================================================

def build(data_dir):
    headers, subtitles, legend_blurbs = load_data(data_dir)
    roadmap = load_roadmap(data_dir)

    wb = Workbook()
    wb.remove(wb.active)

    # SHEET ORDER (controlled by SHEET_ORDER_FIRST / SHEET_ORDER_LAST constants):
    # Legend, To Do, [per-header alphabetical], Pending Dependencies, Coverage Failures.
    sorted_headers = sorted(headers.keys())
    header_blurbs  = [(h, legend_blurbs.get(h, "")) for h in sorted_headers]

    builders = {
        "Legend":               lambda: build_legend_sheet(wb, header_blurbs),
        "To Do":                lambda: build_todo_sheet(wb, headers, roadmap),
        "Pending Dependencies": lambda: build_pending_deps(wb, headers),
        "Coverage Failures":    lambda: build_coverage_failures(wb, headers),
    }

    # Fixed-first sheets
    for name in SHEET_ORDER_FIRST:
        builders[name]()
    # Per-header sheets
    for hdr in sorted_headers:
        build_data_sheet(wb, hdr, subtitles.get(hdr), headers[hdr])
    # Fixed-last sheets
    for name in SHEET_ORDER_LAST:
        builders[name]()

    return wb


if __name__ == "__main__":
    wb = build(DATA_DIR)
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH}")
    print(f"  data:    {DATA_DIR}")
    print(f"  sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
